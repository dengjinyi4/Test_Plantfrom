# encoding=utf-8
import datetime
import sys
from utils.db_info import *
import meitimodle,collections
from openpyxl import  Workbook
from openpyxl import load_workbook
# excel导出
def exportXls(result,day):
    if len(str(result)) > 0 :
        addr = "./abc.xlsx"
        wb = load_workbook(addr)
        sheet = wb.create_sheet(day)
        sheet.append(['媒体id','媒体名称','广告位id','广告位名称','小于5pv个数','大于120pv个数','5120pv个数','小于5uv个数','大于120uv个数','5120uv个数','有效个数','有效时间'])
        for i in result:
            media_id=i[0].get('media_id')
            media_name=i[0].get('media_name')
            adzone_id=i[0].get('adzone_id')
            adzone_name=i[0].get('adzone_name')
            pv5=i[0].get('<5pv')
            pv120=i[0].get('>120pv')
            pv5120=i[0].get('5120pv')
            uv5=i[0].get('<5uv')
            uv120=i[0].get('>120uv')
            uv5120=i[0].get('5120uv')
            copywxuv=str(i[0].get('copywxuv'))
            pvtime=str(i[0].get('pvtime'))
            # sheet.append(result[i])
            sheet.append([media_id,media_name,adzone_id,adzone_name,pv5,pv120,pv5120,uv5,uv120,uv5120,copywxuv,pvtime])
        wb.save("./abc.xlsx")
    return 1
if __name__ == '__main__':
    print 1
    test=meitimodle.media(day='20190822')
    # test.getmediainfor()
    tmpresult=test.get_clickcount()
    # result=meitimodle.media('20190816')
    # tmpresult=result.getmediainfor()
    exportXls(tmpresult,'20190822')

