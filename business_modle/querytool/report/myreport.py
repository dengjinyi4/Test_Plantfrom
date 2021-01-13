# -*- coding: utf-8 -*-
__author__ = 'jinyi'
from business_modle.querytool import db
import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import  Workbook,load_workbook
from flask import current_app
import os

class myreport(object):
    def __init__(self,begintime='',endtime='',adzoneids='',advertiser_id='',actid='',region='',showadzone='',tagorad='',isstatus='',showbaidu='',searchword='',type='',yjf_data_report_id='',value='',update_type=''):
        self.begintime=begintime
        self.endtime=endtime
        self.adzoneids=adzoneids
        self.advertiser_id=advertiser_id
        self.actid=actid
        self.region=region
        self.showadzone=showadzone
        self.showbaidu=showbaidu
        self.searchword=searchword
        self.tagorad=tagorad
        self.isstatus=isstatus
        self.tmpJingJiaMediaID = ''',,3500,3481,3284,3486,3532,3568,3552,3609,3570,3575,3555,3628,3623,3573,3621,3620,3559,3564,3653,3553,3487,2205,3686,3666,3647,3691,3653,3706,3698,3704,3709,7803,3703,3712,3719,3695,3284,3500,3486,3481,3609,3575,3555,3568,3604,3711,3728,3733,3758,3752,3753,3754,3706,,'''
        self.tmpBaiDuAdzoneID = ''',,3171,4736,5302,5303,5402,5403,5908,5909,6329,6476,6677,6705,6742,6749,6810,6910,7221,7223,7344,7360,7396,7459,7482,7491,7492,7493,7508,7510,7566,7604,7605,7607,7608,7609,7610,7611,7612,7613,7768,,'''
        self.tmpQiShiKaAdzoneID = ''',,7309,6694,7587,,'''
        self.tmpTaoLiJinAdzoneID = ''',,6917,7601,3105,7434,7626,7646,7466,7642,7240,7673,7709,7710,7765,7730,,'''
        self.type=type
        self.yjf_data_report_id=yjf_data_report_id
        self.value=value
        self.update_type=update_type

    # 根据输入的日期返回时间列表
    # ['2020-04-01', '2020-04-02', '2020-04-03', '2020-04-04', '2020-04-05', '2020-04-06']
    def gettimelist(self,):
        bgintime=datetime.datetime.strptime(self.begintime,'%Y-%m-%d')
        endtime=datetime.datetime.strptime(self.endtime,'%Y-%m-%d')
        dates=[]
        while bgintime<=endtime:
            dates.append(bgintime.strftime('%Y-%m-%d')[0:10])
            bgintime=bgintime+relativedelta(days=+1)
        return dates




    def getresfloadtoint(self,res):
        '''
        处理fload展示小数据点.0的问题
        '''
        tmpres=[]
        for i in res:
            # print i
            tmplist=list(i)
            for k in range(len(tmplist)):
                if type(tmplist[k])==float:
                    if int(tmplist[k])==tmplist[k]:
                        tmplist[k]=int(tmplist[k])
            tmpres.append(tmplist)
        return tmpres
    # 导出excel
    def exportexcel(self,filed,res,excelname):
        if len(res)<>0:
            res=list(res)
            # res.insert(0,headtr)
            res.insert(0,filed)
            wb=Workbook()
            sheet=wb.active
            for i in range(len(res)):
                sheet.append(res[i])
            try:
                # wb.save("../../../static/result/reportall.xlsx")
                wb.save("./static/result/{0}.xlsx".format(excelname))
            except Exception as e:
                print e.message
        return ''

    def getfieldname(self,field,tmpdit):
        '''
        根据字段名称增加描述，鼠标停留后可见
        '''
        tmp='''<td title="{0}">{1}</td>'''
        tmpv='''<td>{0}</td>'''
        field=[tmp.format(tmpdit[x],x) if x in tmpdit.keys() else tmpv.format(x) for x in field]
        return field


    # 汇总 菜单名---- 毛利表-分媒体毛利
    def getallreport(self):
        daylist=self.gettimelist()

        if self.searchword.replace("竞价","")<>self.searchword:
            tmpwhere=''' and media_id in ({tmpJingJiaMediaID}) '''.format(tmpJingJiaMediaID=self.tmpJingJiaMediaID.replace(",,",""))
        elif self.searchword.replace("百度","")<>self.searchword:
            tmpwhere=''' and adzone_id in ({tmpBaiDuAdzoneID}) '''.format(tmpBaiDuAdzoneID=self.tmpBaiDuAdzoneID.replace(",,",""))
        elif self.searchword.replace("骑士","")<>self.searchword:
            tmpwhere=''' and adzone_id in ({tmpQiShiKaAdzoneID}) '''.format(tmpQiShiKaAdzoneID=self.tmpQiShiKaAdzoneID.replace(",,",""))
        elif self.searchword.replace("淘礼金","")<>self.searchword:
            tmpwhere=''' and adzone_id in ({tmpTaoLiJinAdzoneID}) '''.format(tmpTaoLiJinAdzoneID=self.tmpTaoLiJinAdzoneID.replace(",,",""))
        else:
            tmpwhere=''


        tmpsql='''select adzone_id as ID, max(adzone_name) as 广告位, ifnull(max(settle_type),"") as 结算
                    ,ifnull((case when  instr(max(adzone_name),"亿起发")>0 and instr(max(adzone_name),"jinlika")<=0 and instr(max(adzone_name),"省点")<=0 then "--" when sum(jf97_consume)>0 then  "R5" else "R1" end),"") as R级别
                    ,concat(round(ifnull(sum(jf97_consume)/sum(consume),0)*100,0),"%") as 加粉比例'''

        tmpsql0 = ''',round(ifnull(sum(consume),0),0) as 总消耗
                    ,round(ifnull(sum(cash_consume)-sum(linkage_cash_consume),0),0) as 去联现金
                    ,round(ifnull(sum(platform_profit)-sum(linkage_cash_consume),0),0) as 去联毛利
                    ,concat(round(ifnull((sum(platform_profit)-sum(linkage_cash_consume))/(sum(cash_consume)-sum(linkage_cash_consume)),0)*100,0),"%") as 去联毛利率'''
        tmp1=''
        tmp11=''
        tmp12=''
        colspanx=0
        for i in daylist:  #消耗
            tmp12=tmp12+''', round(ifnull(sum(case date when "{0}" then adzone_click else 0 end),0),0) as "{1}" '''.format(i,i.replace("2020-","")) #入口点击
            tmp1=tmp1+''', round(ifnull(sum(case date when "{0}" then consume else 0 end),0),0) as "{1}" '''.format(i,i.replace("2020-",""))  #消耗
            tmp11=tmp11+''', round(ifnull(sum(case date when "{0}" then consume-linkage_consume else 0 end),0),0) as "{1}" '''.format(i,i.replace("2020-",""))  #去联消耗
            colspanx=colspanx+1
            # tmpsql=tmpsql+tmp1
        tmp2=''
        for i in daylist: #去除联动平台毛利
            tmp2=tmp2+''',round(ifnull(sum(case date when "{0}" then platform_profit-linkage_cash_consume else 0 end),0),0) as "{1}" '''.format(i,i.replace("2020-",""))
        tmp3=''
        tmp31 = ''
        for i in daylist: #入口点击成本
            tmp3=tmp3+''',round(ifnull(sum(case date when "{0}" then adzone_ecpc else 0 end),0),2) as "{1}"'''.format(i,i.replace("2020-",""))
            tmp31=tmp31+''',round(ifnull(sum(case date when "{0}" then adzone_cost else 0 end),0),0) as "{1}"'''.format(i,i.replace("2020-",""))  #成本
        tmp4=''
        tmp41=''
        for i in daylist: #去除联动毛利率
            tmp4=tmp4+''',concat(round(ifnull(sum(case when date="{0}" and (cash_consume-linkage_cash_consume)>0 then (platform_profit-linkage_cash_consume)  else 0 end)/sum(case when date="{1}" and (cash_consume-linkage_cash_consume)>0 then (cash_consume-linkage_cash_consume) else 0 end),0)*100,0),"%") as "{2}"'''.format(i,i,i.replace("2020-",""))
            tmp41=tmp41+''',round(ifnull(sum(case date when "{0}" then consume-linkage_consume else 0 end)/sum(case date when "{0}" then ad_click else 0 end),0),2) as "{1}"'''.format(i,i.replace("2020-",""))  #CPC
        tmpend='''      ,max(case
                        when instr("{tmpJingJiaMediaID}",concat(",",media_id,","))>0  then "竞价媒体"
                        when instr("{tmpBaiDuAdzoneID}",concat(",",adzone_id,","))>0  then "百度广告位"
                        when instr("{tmpQiShiKaAdzoneID}",concat(",",adzone_id,","))>0  then "骑士卡广告位"
                        when instr("{tmpTaoLiJinAdzoneID}",concat(",",adzone_id,","))>0  then "淘礼金广告位"
                        else 0 end) as 分类型
                    from tt.tt_adzone_data
                where date between "{begin1}" and "{end1}"
                '''.format(tmpJingJiaMediaID=self.tmpJingJiaMediaID,tmpBaiDuAdzoneID=self.tmpBaiDuAdzoneID,tmpQiShiKaAdzoneID=self.tmpQiShiKaAdzoneID,tmpTaoLiJinAdzoneID=self.tmpTaoLiJinAdzoneID,begin1=self.begintime,end1=self.endtime)
        tmpgroup =''' group by adzone_id ;'''
        tmpsql=tmpsql+tmpsql0+tmp12+tmp1+tmp11+tmp2+tmp3+tmp31+tmp4+tmp41+tmpend+tmpwhere+tmpgroup
        headtr='''<tr><td>&nbsp;</td><td>&nbsp;</td><td>&nbsp;</td><td>&nbsp;</td><td>&nbsp;</td><td>&nbsp;</td><td>&nbsp;</td><td>&nbsp;</td><td>&nbsp;</td>
            <td align="center" colspan={0}>入口点击</td><td align="center" colspan={0}>消耗</td><td align="center" colspan={0}>去联消耗</td><td align="center" colspan={1}>去除联动平台毛利</td><td align="center" colspan={2}>入口点击成本</td><td align="center" colspan={2}>成本</td><td align="center" colspan={3}>去除联动毛利率</td><td align="center" colspan={3}>CPC</td><td>分类型</td></tr>'''.format(colspanx,colspanx,colspanx,colspanx)
        # print tmpsql
        res,filed=db.selectsqlnew('devtidb',tmpsql)
        res=self.getresfloadtoint(res)
        self.exportexcel(filed,res,"reportall")

        tmpsqlsum = 'select "","","","","" ' + tmpsql0 + tmp12 + tmp1 + tmp11 + tmp2 + tmp3 + tmp31 + tmp4 + tmp41 + tmpend+tmpwhere
        ressum,filedsum=db.selectsqlnew('devtidb',tmpsqlsum)
        ressum=self.getresfloadtoint(ressum)

        return res,filed,tmpsql ,headtr,ressum



    #广告主联动毛利    菜单名---- 毛利表-联动客户毛利
    def getadvliandong(self):
        tmpsql='''
        select addata.date as 日期
            ,round(ifnull(sum(addata.linkage_consume),0),0) as 联动总消耗
            ,round(ifnull(sum(addata.linkage_cash_consume),0),0) as 联动现金消耗
            ,concat(round(ifnull(sum(addata.linkage_cash_consume),0)/ifnull(sum(addata.linkage_consume),0)*100,0),"%") as 联动现金占比
            ,round(ifnull(sum(adzonedata.adzone_cost),0),0) as 联动成本
            ,round(ifnull(sum(addata.linkage_cash_consume)-sum(adzonedata.adzone_cost),0),0) as 联动毛利
            ,round(ifnull(sum(addata.consume),0),0) as 联动广告主总消耗
            ,concat(round(ifnull(sum(addata.linkage_consume)/sum(addata.consume),0)*100,0),"%") as 联动消耗占比
            ,round(ifnull(sum(case addata.advertiser_id when  5209 then addata.linkage_consume  else 0 end),0),0) as 联动消耗
            ,round(ifnull(sum(case addata.advertiser_id when  5209 then addata.linkage_cash_consume  else 0 end),0),0) as 联动现金消耗
            ,concat(round(ifnull(sum(case addata.advertiser_id when  5209 then addata.linkage_cash_consume  else 0 end),0)/ifnull(sum(case addata.advertiser_id when  5209 then addata.linkage_consume  else 0 end),0)*100,0),"%") as 联动现金比例
            ,round(ifnull(sum(case addata.advertiser_id when  5209 then adzonedata.adzone_cost  else 0 end),0),0) as 联动成本
            ,round(ifnull(sum(case addata.advertiser_id when  5209 then addata.consume  else 0 end),0),0) as 总消耗
            ,concat(round(ifnull(sum(case addata.advertiser_id when  5209 then addata.linkage_consume  else 0 end)/sum(case addata.advertiser_id when  5209 then addata.consume  else 0 end),0)*100,0),"%") as 联动消耗占比
            ,round(ifnull(sum(case addata.advertiser_id when  5209 then e.e  else 0 end),0),0) as 联动效果数
            ,round(ifnull(sum(case addata.advertiser_id when  5209 then addata.linkage_consume  else 0 end),0)/ifnull(sum(case addata.advertiser_id when  5209 then e.e  else 0 end),0),0) as 总消耗_转化成本
            ,round(ifnull(sum(case addata.advertiser_id when  5209 then adzonedata.adzone_cost  else 0 end),0)/ifnull(sum(case addata.advertiser_id when  5209 then e.e  else 0 end),0),0) as 成本_转化成本
            ,round(ifnull(sum(case addata.advertiser_id when  5002 then addata.linkage_consume  else 0 end),0),0) as 联动消耗
            ,round(ifnull(sum(case addata.advertiser_id when  5002 then addata.linkage_cash_consume  else 0 end),0),0) as 联动现金消耗
            ,concat(round(ifnull(sum(case addata.advertiser_id when  5002 then addata.linkage_cash_consume  else 0 end),0)/ifnull(sum(case addata.advertiser_id when  5002 then addata.linkage_consume  else 0 end),0)*100,0),"%") as 联动现金占比
            ,round(ifnull(sum(case addata.advertiser_id when  5002 then adzonedata.adzone_cost  else 0 end),0),0) as 联动成本
            ,round(ifnull(sum(case addata.advertiser_id when  5002 then addata.consume  else 0 end),0),0) as 总消耗
            ,concat(round(ifnull(sum(case addata.advertiser_id when  5002 then addata.linkage_consume  else 0 end)/sum(case addata.advertiser_id when  5002 then addata.consume  else 0 end),0)*100,0),"%") as 联动消耗占比
            ,round(ifnull(sum(case addata.advertiser_id when  5002 then e.e  else 0 end),0),0) as 联动效果数
            ,round(ifnull(sum(case addata.advertiser_id when  5002 then addata.linkage_consume  else 0 end),0)/ifnull(sum(case addata.advertiser_id when  5002 then e.e  else 0 end),0),0) as 总消耗_转化成本
            ,round(ifnull(sum(case addata.advertiser_id when  5002 then adzonedata.adzone_cost  else 0 end),0)/ifnull(sum(case addata.advertiser_id when  5002 then e.e  else 0 end),0),0) as 成本_转化成本
        from (    
            select date ,advertiser_id ,linkage_consume ,linkage_cash_consume,consume from tt.tt_advertiser_data    
            where advertiser_id in (5209,5002) and date between "{begin1}" and "{end1}"
        ) addata left join ( 
            select a.date,a.advertiser_id,sum(adzone_cost) as adzone_cost 
            from( 
                select sum(adzone_cost) as adzone_cost, date  , (case when instr(max(adzone_name), "anyi") > 0 then 5209 when instr(max(adzone_name), "马上") > 0 then 5002  else 0 end)     as advertiser_id
                from tt.tt_adzone_data 
                where adzone_name like "%亿起发%" and date between "{begin2}" and "{end2}"
                group by date,adzone_name 
            ) a group by a.date,a.advertiser_id
        ) adzonedata on addata.date=adzonedata.date and addata.advertiser_id=adzonedata.advertiser_id
         left join (select e.advertiser_id, e.date,
                          sum(ELT(a.effect_type, t1_num, t2_num,t3_num,t4_num,t5_num,t6_num,t7_num,t8_num,t9_num,t10_num,t11_num,t12_num,t13_num,t14_num,t15_num,t16_num)) as e
                   from tt.tt_advertiser_linkage_data e left join tt.tt_advertiser_effect_standard a on e.advertiser_id=a.advertiser_id
                   where e.date  between "{begin3}" and "{end3}"
                group by e.advertiser_id,e.date
            ) e  on e.advertiser_id=addata.advertiser_id and e.date=addata.date
        group by addata.date;
        '''.format(begin1=self.begintime,end1=self.endtime,begin2=self.begintime,end2=self.endtime,begin3=self.begintime,end3=self.endtime)
        print tmpsql
        headtr = '''<tr><td colspan=8>&nbsp;</td>
                    <td align="center" colspan=9>5209anyihua</td>
                    <td align="center" colspan=9>5002马上消费金融</td>
                    </tr>'''
        res,filed=db.selectsqlnew('devtidb',tmpsql)
        res=self.getresfloadtoint(res)
        self.exportexcel(filed,res,"reportadvliandong")

        return res,filed,tmpsql,headtr

    # 汇总   菜单名---- 毛利表-平台毛利细化
    def getptmaoli(self):

        tmpsql = '''select adzonedata.date as 日期
                        ,round(ifnull(adzonedata.现金消耗,0),0) as 现金消耗
                        ,round(ifnull(adzonedata.媒体成本,0),0) as 媒体成本
                        ,round(ifnull(adzonedata.平台毛利,0),0) as 平台毛利
                        ,concat(round(ifnull(adzonedata.平台毛利/adzonedata.现金消耗,0)*100,0),"%") as 毛利率
                        ,round(ifnull(adzonedata.媒购去联动现金消耗,0),0) as 媒购去联动现金消耗
                        ,round(ifnull(adzonedata.媒购媒体成本,0),0) as 媒购媒体成本
                        ,round(ifnull(adzonedata.媒购去联动毛利,0),0) as 媒购去联动毛利
                        ,concat(round(ifnull((adzonedata.媒购去联动毛利)/(adzonedata.媒购去联动现金消耗),0)*100,0),"%") as 毛利率
                        ,round(ifnull(adzonedata.CPS去联动现金消耗,0),0) as CPS去联动现金消耗
                        ,round(ifnull(adzonedata.CPS媒体成本,0),0) as CPS媒体成本
                        ,round(ifnull(adzonedata.CPS去联动毛利,0),0) as CPS去联动毛利
                        ,concat(round(ifnull((adzonedata.CPS去联动毛利)/(adzonedata.CPS去联动现金消耗),0)*100,0),"%") as 毛利率
                        ,round(ifnull(adzonedata.竞价平台去联现金,0),0) as 竞价平台去联现金
                        ,round(ifnull(adzonedata.竞价平台成本,0),0) as 竞价平台成本
                        ,round(ifnull(adzonedata.竞价平台去联现金-adzonedata.竞价平台成本,0),0) as 竞价平台毛利
                        ,concat(round(ifnull((adzonedata.竞价平台去联现金-adzonedata.竞价平台成本)/(adzonedata.竞价平台去联现金),0)*100,0),"%") as 毛利率
                        ,round(ifnull(adzonedata.百度媒体去联现金,0),0) as 百度媒体去联现金
                        ,round(ifnull(adzonedata.百度媒体成本,0),0) as 百度媒体成本
                        ,round(ifnull(adzonedata.百度媒体去联现金-adzonedata.百度媒体成本,0),0) as 百度媒体毛利
                        ,concat(round(ifnull((adzonedata.百度媒体去联现金-adzonedata.百度媒体成本)/(adzonedata.百度媒体去联现金),0)*100,0),"%") as 毛利率
                        ,round(ifnull(adzonedata.骑士卡媒体去联现金,0),0) as 骑士卡媒体去联现金
                        ,round(ifnull(adzonedata.骑士卡媒体成本,0),0) as 骑士卡媒体成本
                        ,round(ifnull(adzonedata.骑士卡媒体去联现金-adzonedata.骑士卡媒体成本,0),0) as 骑士卡媒体毛利
                        ,concat(round(ifnull((adzonedata.骑士卡媒体去联现金-adzonedata.骑士卡媒体成本)/(adzonedata.骑士卡媒体去联现金),0)*100,0),"%") as 毛利率
                        ,round(ifnull(adzonedata.淘礼金媒体去联现金,0),0) as 淘礼金媒体去联现金
                        ,round(ifnull(adzonedata.淘礼金媒体成本,0),0) as 淘礼金媒体成本
                        ,round(ifnull(adzonedata.淘礼金媒体去联现金-adzonedata.淘礼金媒体成本,0),0) as 淘礼金媒体毛利
                        ,concat(round(ifnull((adzonedata.淘礼金媒体去联现金-adzonedata.淘礼金媒体成本)/(adzonedata.淘礼金媒体去联现金),0)*100,0),"%") as 毛利率
                        ,round(ifnull(addata.联动现金消耗+addata.还呗5213联动现金消耗,0),0) as 联动现金消耗
                        ,round(ifnull(adzonedata.联动媒体成本,0),0) as 联动媒体成本
                        ,round(ifnull(addata.联动现金消耗+addata.还呗5213联动现金消耗-adzonedata.联动媒体成本,0),0) as 联动毛利
                        ,concat(round(ifnull((addata.联动现金消耗+addata.还呗5213联动现金消耗-adzonedata.联动媒体成本)/(addata.联动现金消耗+addata.还呗5213联动现金消耗),0)*100,0),"%") as 毛利率
                        ,round(ifnull(addata.广点通现金消耗,0),0) as 广点通现金消耗
                        from (
                            select a.date
                             ,sum(a.ad_show) as 广告曝光
                             ,sum(a.cash_consume) as 现金消耗
                             ,sum(a.adzone_cost) as 媒体成本
                             ,sum(platform_profit) as 平台毛利
                             ,sum(case  when (instr(adzone_name,"亿起发")>0 and instr(adzone_name,"jinlika")<=0 and instr(adzone_name,"省点")<=0 and instr(adzone_name,"优惠现报")<=0) or (instr("{tmpJingJiaMediaID}",concat(",",media_id,","))>0) or (instr("{tmpBaiDuAdzoneID}{tmpQiShiKaAdzoneID}{tmpTaoLiJinAdzoneID}",concat(",",a.adzone_id,","))>0) then 0 when lower(settle_type)="cps" then 0 else ifnull(a.cash_consume,0)-ifnull(linkage_cash_consume,0) end) as 媒购去联动现金消耗
                             ,sum(case  when (instr(adzone_name,"亿起发")>0 and instr(adzone_name,"jinlika")<=0 and instr(adzone_name,"省点")<=0 and instr(adzone_name,"优惠现报")<=0) or (instr("{tmpJingJiaMediaID}",concat(",",media_id,","))>0) or (instr("{tmpBaiDuAdzoneID}{tmpQiShiKaAdzoneID}{tmpTaoLiJinAdzoneID}",concat(",",a.adzone_id,","))>0) then 0 when lower(settle_type)="cps" then 0 else ifnull(adzone_cost,0) end) as 媒购媒体成本
                             ,sum(case  when (instr(adzone_name,"亿起发")>0 and instr(adzone_name,"jinlika")<=0 and instr(adzone_name,"省点")<=0 and instr(adzone_name,"优惠现报")<=0) or (instr("{tmpJingJiaMediaID}",concat(",",media_id,","))>0) or (instr("{tmpBaiDuAdzoneID}{tmpQiShiKaAdzoneID}{tmpTaoLiJinAdzoneID}",concat(",",a.adzone_id,","))>0) then 0 when lower(settle_type)="cps" then 0 else ifnull(platform_profit,0)-ifnull(linkage_cash_consume,0) end) as 媒购去联动毛利
                             ,sum(case  when (instr(adzone_name,"亿起发")>0 and instr(adzone_name,"jinlika")<=0 and instr(adzone_name,"省点")<=0 and instr(adzone_name,"优惠现报")<=0) or (instr("{tmpJingJiaMediaID}",concat(",",media_id,","))>0) or (instr("{tmpBaiDuAdzoneID}{tmpQiShiKaAdzoneID}{tmpTaoLiJinAdzoneID}",concat(",",a.adzone_id,","))>0) then 0 when lower(settle_type)="cps" then ifnull(a.cash_consume,0)-ifnull(linkage_cash_consume,0) else 0 end) as CPS去联动现金消耗
                             ,sum(case  when (instr(adzone_name,"亿起发")>0 and instr(adzone_name,"jinlika")<=0 and instr(adzone_name,"省点")<=0 and instr(adzone_name,"优惠现报")<=0) or (instr("{tmpJingJiaMediaID}",concat(",",media_id,","))>0) or (instr("{tmpBaiDuAdzoneID}{tmpQiShiKaAdzoneID}{tmpTaoLiJinAdzoneID}",concat(",",a.adzone_id,","))>0) then 0 when lower(settle_type)="cps" then ifnull(adzone_cost,0) else 0 end) as CPS媒体成本
                             ,sum(case  when (instr(adzone_name,"亿起发")>0 and instr(adzone_name,"jinlika")<=0 and instr(adzone_name,"省点")<=0 and instr(adzone_name,"优惠现报")<=0) or (instr("{tmpJingJiaMediaID}",concat(",",media_id,","))>0) or (instr("{tmpBaiDuAdzoneID}{tmpQiShiKaAdzoneID}{tmpTaoLiJinAdzoneID}",concat(",",a.adzone_id,","))>0) then 0 when lower(settle_type)="cps" then ifnull(platform_profit,0)-ifnull(linkage_cash_consume,0) else 0 end) as CPS去联动毛利
                             ,sum(case  when instr(adzone_name,"亿起发")>0 and instr(adzone_name,"jinlika")<=0 and instr(adzone_name,"省点")<=0 and instr(adzone_name,"优惠现报")<=0 then ifnull(adzone_cost,0) else 0 end) as 联动媒体成本
                             ,sum(case  when instr("{tmpJingJiaMediaID}",concat(",",media_id,","))>0  then ifnull(a.cash_consume,0)-ifnull(linkage_cash_consume,0) else 0 end) as 竞价平台去联现金
                             ,sum(case  when instr("{tmpJingJiaMediaID}",concat(",",media_id,","))>0  then ifnull(adzone_cost,0) else 0 end) as 竞价平台成本
                             ,sum(case  when instr("{tmpBaiDuAdzoneID}",concat(",",a.adzone_id,","))>0  then ifnull(a.cash_consume,0)-ifnull(linkage_cash_consume,0) else 0 end) as 百度媒体去联现金
                             ,sum(case  when instr("{tmpBaiDuAdzoneID}",concat(",",a.adzone_id,","))>0  then ifnull(adzone_cost,0) else 0 end) as 百度媒体成本
                             ,sum(case  when instr("{tmpQiShiKaAdzoneID}",concat(",",a.adzone_id,","))>0  then ifnull(a.cash_consume,0)-ifnull(linkage_cash_consume,0) else 0 end) as 骑士卡媒体去联现金
                             ,sum(case  when instr("{tmpQiShiKaAdzoneID}",concat(",",a.adzone_id,","))>0  then ifnull(adzone_cost,0) else 0 end) as 骑士卡媒体成本
                             ,sum(case  when instr("{tmpTaoLiJinAdzoneID}",concat(",",a.adzone_id,","))>0  then ifnull(a.cash_consume,0)-ifnull(linkage_cash_consume,0) else 0 end) as 淘礼金媒体去联现金
                             ,sum(case  when instr("{tmpTaoLiJinAdzoneID}",concat(",",a.adzone_id,","))>0  then ifnull(adzone_cost,0) else 0 end) as 淘礼金媒体成本
                             ,sum(b.百度现金消耗) as 百度现金消耗
                             ,sum(a.adzone_cost/a.ad_show*b.百度曝光) as 百度成本
                             ,sum(b.百度现金消耗)-sum(a.adzone_cost/a.ad_show*b.百度曝光) as 百度毛利
                            from tt.tt_adzone_data a
                                left join (
                                    select date ,adzone_id,sum(ad_show) as ad_show,sum(cash_consume) as cash_consume
                                        ,sum(case when advertiser_id=3051 or advertiser_id=3582 or advertiser_id=6402 or advertiser_id=6401 then ad_show else 0 end) as 百度曝光
                                        ,sum(case when advertiser_id=3051 or advertiser_id=3582 or advertiser_id=6402 or advertiser_id=6401 then cash_consume else 0 end) as 百度现金消耗
                                    from tt.tt_advertiser_adzone
                                    where date between "{begin1}" and "{end1}"
                                    group by date,adzone_id
                                ) b on a.date=b.date and a.adzone_id=b.adzone_id
                            where a.date between "{begin1}" and "{end1}"
                            group by a.date
                        ) adzonedata left join(    
                            select date        
                            ,sum(case  when advertiser_id=5213 then cash_consume else 0 end) as 还呗5213联动现金消耗
                            ,sum(case  when advertiser_id=3051 or  advertiser_id=3582 then cash_consume else 0 end) as 百度现金消耗 
                            ,sum(case  when instr(advertiser_name,"广点通")>0 then 0 else linkage_cash_consume end) as 联动现金消耗
                            ,sum(case  when instr(advertiser_name,"广点通")>0 then cash_consume else 0 end) as 广点通现金消耗
                            from tt.tt_advertiser_data    where 1=1 
                            and date between "{begin2}" and "{end2}"    group by date
                        ) addata on adzonedata.date=addata.date;'''.format(begin1=self.begintime, end1=self.endtime,
                                                                           begin2=self.begintime,
                                                                           end2=self.endtime,tmpJingJiaMediaID=self.tmpJingJiaMediaID,tmpBaiDuAdzoneID=self.tmpBaiDuAdzoneID,tmpQiShiKaAdzoneID=self.tmpQiShiKaAdzoneID,tmpTaoLiJinAdzoneID=self.tmpTaoLiJinAdzoneID)
        print tmpsql
        res, filed = db.selectsqlnew('devtidb', tmpsql)
        res = self.getresfloadtoint(res)
        self.exportexcel(filed,res,"reportptmaoli")

        tmpsqlsum = tmpsql.replace("group by a.date","").replace("select a.date","select 1 as date ").replace("select adzonedata.date","select '汇总' ")
        ressum, filedsum = db.selectsqlnew('devtidb', tmpsqlsum)
        ressum = self.getresfloadtoint(ressum)


        return res, filed, tmpsql+'''|||'''+tmpsqlsum ,ressum

        # 菜单名：毛利细化-广告主维度
        # 查询日期，广告位（广告位为可选）
    def getreportptmaoliadtag(self):
        daylist = self.gettimelist()

        tmpad = ''
        tmpadz = ''
        if self.advertiser_id <> '' and self.advertiser_id <> 'mytest':
            tmpad = ''' and aa.advertiser_id in ({0}) '''.format(self.advertiser_id)
        else:
            tmpad = ' '

        if self.adzoneids <> '' and self.adzoneids <> 'mytest':
            tmpadz = ''' and aa.adzone_id in ({0}) '''.format(self.adzoneids)
        else:
            tmpadz = ''

        tmpshowadzone = ''
        tmpshowadzone1 = ''
        tmpshowadzonesum = ''
        if self.showadzone == '1':
            tmpshowadzone = ''' ,aa.adzone_id as  广告位ID,aa.adzone_name as 广告位名称 '''
            tmpshowadzone1 = ''' ,aa.adzone_id '''
            tmpshowadzonesum = ''' ,"--" as  广告位ID,"--" as 广告位名称 '''
        else:
            tmpshowadzone = ''
            tmpshowadzone1 = ''
            tmpshowadzonesum = ''

        tmptagorad = ''
        tmptagorad1 = ''
        tmptagoradsum = ''
        tmpCB = ''
        tmpCBsum = ''
        tmpCBTable = ''
        if self.tagorad == 'tag':
            tmptagorad = ''' ifnull(tag, "")    as 类型 '''
            tmptagorad1 = ''' aa.tag '''
            tmptagoradsum = ''' "--" '''
        elif self.tagorad == 'ad':
            tmptagorad = '''  ifnull(tag, "")    as 类型 ,aa.advertiser_id  as 广告主ID, aa.advertiser_name  as 广告主名称 '''
            tmptagorad1 = ''' aa.advertiser_id '''
            tmptagoradsum = ''' "--","--","--" '''
            tmpCB = '''  ,round(ifnull((sum(aa.consume) - sum(aa.linkage_consume)) / sum(
                                ELT(c.effect_type, aa.t1_num, aa.t2_num, aa.t3_num, aa.t4_num, aa.t5_num, aa.t6_num, aa.t7_num, aa.t8_num, aa.t9_num, aa.t10_num, aa.t11_num,
                            aa.t12_num, aa.t13_num, aa.t14_num, aa.t15_num, aa.t16_num)), 0), 2)    as 转化成本,ifnull(c.standard,"") as 要求 '''
            tmpCBTable = ''' left join tt.tt_advertiser_effect_standard c on aa.advertiser_id = c.advertiser_id  '''
            tmpCBsum = ''' ,"--","--" '''
        else:
            tmptagorad = ''
            tmptagorad1 = ''
            tmpCB = ''
            tmpCBTable = ''

        colspanx = 0
        tmp1 = ''
        tmpMLL = ''
        tmp1CB = ''
        for i in daylist:  # 转化成本
            colspanx = colspanx + 1
            tmp1 = tmp1 + ''' , round(ifnull(sum(case aa.date when "{begin1}" then aa.cash_consume-aa.linkage_cash_consume-ifnull(b.singleshowvalue,0)*ifnull(aa.ad_show,0) else 0 end), 0), 0) as "{begin2}" '''.format(
                begin1=i, begin2=i.replace("2020-", ""))
            tmpMLL = tmpMLL + ''' , concat(round(ifnull(
                                    (sum(case aa.date when "{begin1}" then aa.cash_consume-aa.linkage_cash_consume else 0 end)
                                      - sum(case aa.date when "{begin2}" then ifnull(b.singleshowvalue,0)*ifnull(aa.ad_show,0) else 0 end))
                                    /sum(case aa.date when "{begin3}" then aa.cash_consume-aa.linkage_cash_consume else 0 end)   
                                ,0)*100 , 0),"%") as "{begin4}" '''.format(
                begin1=i, begin2=i, begin3=i, begin4=i.replace("2020-", "")+"毛利率")
            tmp1CB = tmp1CB + ''' ,round(ifnull(sum(case aa.date when "{begin1}" then aa.consume - aa.linkage_consume else 0 end) / sum(case aa.date when "{begin1}" then ELT(c.effect_type, aa.t1_num, aa.t2_num, aa.t3_num, aa.t4_num, aa.t5_num, aa.t6_num, aa.t7_num, aa.t8_num, aa.t9_num, aa.t10_num, aa.t11_num,
                            aa.t12_num, aa.t13_num, aa.t14_num, aa.t15_num, aa.t16_num)  else 0 end), 0), 2)  as "转化成本{begin2}" '''.format(
                begin1=i, begin2=i.replace("2020-", ""))

        if self.tagorad == 'ad':
            tmp1CB = tmp1CB
        else:
            tmp1CB = ''

        tmp = '''select  {tmptagorad}  {tmpshowadzone} 
                 , round(ifnull(sum(aa.ad_show), 0), 0)                                                               as 曝光
                 , round(ifnull(sum(aa.ad_click), 0), 0)                                                              as 点击
                 , round(ifnull(sum(aa.consume) - sum(aa.linkage_consume), 0), 0)                                        as 消耗
                 , round(ifnull(sum(aa.cash_consume) - sum(aa.linkage_cash_consume), 0), 0)                                        as 去联现金
                 , concat(round(ifnull(sum(aa.cash_consume) - sum(aa.linkage_cash_consume), 0)/ifnull(sum(aa.consume) - sum(aa.linkage_consume), 0)*100, 0) ,"%")                                       as 现金占比
                 , round(ifnull(sum(aa.cash_consume-aa.linkage_cash_consume),0)-sum(ifnull(b.singleshowvalue,0)*ifnull(aa.ad_show,0)),0) as 总盈亏
                 , concat(round((ifnull(sum(aa.cash_consume-aa.linkage_cash_consume),0)-sum(ifnull(b.singleshowvalue,0)*ifnull(aa.ad_show,0)))/(ifnull(sum(aa.cash_consume) - sum(aa.linkage_cash_consume), 0))*100,0),"%")  as 毛利率
                 , round(ifnull((sum(aa.consume) - sum(aa.linkage_consume)) / sum(aa.ad_click), 0), 2)       as CPC
                 , concat(round(ifnull(sum(aa.ad_click) / sum(aa.ad_show) *100 , 0), 0),"%")       as CTR
                 , round(ifnull((sum(aa.consume) - sum(aa.linkage_consume)) / sum(aa.ad_show) * 1000, 0), 0)       as CPM
                 {tmpCB}  {tmp1} {tmpMLL} {tmp1CB}
                 ''' + '''
            from tt.tt_advertiser_adzone aa
                     left join (select adzone_id,date,adzone_cost/ad_show as singleshowvalue,ad_show from tt.tt_adzone_data where adzone_name not like "%亿起发%" or (adzone_name  like "%亿起发%" and adzone_name not like "%anyi%" and adzone_name not like "%马上%")) b on aa.adzone_id = b.adzone_id and aa.date=b.date
                    {tmpCBTable}
            where aa.date between "{begin}" and "{end}" '''.format(begin=self.begintime, end=self.endtime,
                                                                   tmpCBTable=tmpCBTable) + tmpad + tmpadz + '''
            group by   {tmptagorad1}  {tmpshowadzone1} '''

        tmpsql = tmp.format(tmptagorad=tmptagorad, tmpshowadzone=tmpshowadzone, tmptagorad1=tmptagorad1,
                            tmpshowadzone1=tmpshowadzone1,tmpCB=tmpCB,tmp1=tmp1,tmpMLL=tmpMLL,tmp1CB=tmp1CB)

        res, filed = db.selectsqlnew('devtidb', tmpsql)
        res = self.getresfloadtoint(res)
        self.exportexcel(filed, res, "reportptmaoliadtag")

        tmpsqlsum = tmp.format(tmptagorad=tmptagoradsum, tmpshowadzone=tmpshowadzonesum, tmptagorad1='',
                               tmpshowadzone1='',tmpCB=tmpCBsum,tmp1=tmp1,tmpMLL=tmpMLL,tmp1CB=" ").replace("group by", "")
        ressum, filedsum = db.selectsqlnew('devtidb', tmpsqlsum)
        ressum = self.getresfloadtoint(ressum)

        return res, filed, tmpsql, colspanx, ressum

    #菜单名：媒体效果-评估-日表   曝光点击无联动值；消耗包含联动
    # 查询日期，广告位（广告位为可选）
    def getmtpinggu(self):
        daylist=self.gettimelist()

        tmpt01=''
        tmpt02=''
        tmpt03=''
        tmpt01sum=''
        UnmatchedHead=''
        if self.advertiser_id<>'' and self.advertiser_id<>'mytest':
            tmpt01=''' aa.advertiser_id as 广告主ID ,aa.advertiser_name as 广告主名称,aa.adzone_id as 广告位ID ,concat(aa.adzone_name,"&nbsp;-&nbsp;",tad.settle_type)  as 广告位名称 '''
            tmpt01sum = ''' "","","","汇总" '''
            UnmatchedHead = ''' "","","","未匹配效果数" '''
            tmpt02=''' and aa.advertiser_id in ({0}) '''.format(self.advertiser_id)
            tmpt03=''' ,aa.adzone_id '''
        else:
            tmpt01=''' aa.advertiser_id as 广告主ID ,aa.advertiser_name as 广告主名称 '''
            tmpt01sum=''' "","汇总" '''
            UnmatchedHead = ''' "","未匹配效果数" '''

        tmpsql000=''' ,ifnull(b.effect_typename,"") as 考核 ,ifnull(b.standard,"") as 要求 ,ifnull(tag,"") as 类型 '''
        tmp='''select '''+tmpt01+tmpsql000+'''
            ,round(ifnull(sum(ad_show),0),0) as 曝光
            ,round(ifnull(sum(ad_click),0),0) as 点击
            ,round(ifnull(sum(consume)-sum(linkage_consume),0),0) as 消耗
            ,round(ifnull((sum(consume)-sum(linkage_consume))/sum(ad_show)*1000,0),0) as CPM
            ,round(ifnull((sum(consume)-sum(linkage_consume))/sum(ad_click),0),2) as CPC
            ,ifnull(sum(ELT(b.effect_type, t1_num, t2_num,t3_num,t4_num,t5_num,t6_num,t7_num,t8_num,t9_num,t10_num,t11_num,t12_num,t13_num,t14_num,t15_num,t16_num)),0) as 效果数
            ,concat(round(ifnull(sum(ELT(b.effect_type, t1_num, t2_num,t3_num,t4_num,t5_num,t6_num,t7_num,t8_num,t9_num,t10_num,t11_num,t12_num,t13_num,t14_num,t15_num,t16_num))/sum(ad_click),0)*100,0),"%") as 转化率
            ,round(ifnull((sum(consume)-sum(linkage_consume))/sum(ELT(b.effect_type, t1_num, t2_num,t3_num,t4_num,t5_num,t6_num,t7_num,t8_num,t9_num,t10_num,t11_num,t12_num,t13_num,t14_num,t15_num,t16_num)),0),2) as 转化成本
            '''
        tmp1=''
        UnmatchedEffectNum = ''
        UnmatchedEffectNum1 = ''
        colspanx = 0
        for i in daylist: #转化成本
            colspanx=colspanx+1
            tmp1=tmp1+''' ,round(ifnull(sum(case date when "{begin1}" then consume-linkage_consume else 0 end)/
     sum(case date
            when "{begin2}" then ELT(b.effect_type, t1_num, t2_num,t3_num,t4_num,t5_num,t6_num,t7_num,t8_num,t9_num,t10_num,t11_num,t12_num,t13_num,t14_num,t15_num,t16_num)
        else 0 end),0),2) as "{begin3}"'''.format(begin1=i,begin2=i,begin3=i.replace("2020-",""))
            UnmatchedEffectNum = UnmatchedEffectNum + ''' ,sum(case date_format(create_time,"%Y-%m-%d") when "{begin1}" then 1 else 0 end) as "{begin2}" '''.format(begin1=i,begin2=i.replace("2020-",""))
            UnmatchedEffectNum1 = UnmatchedEffectNum1 + ''' ,"" '''
        tmp2=''
        for i in daylist:   #CPC
            tmp2=tmp2+''' , round(ifnull(sum(case date when "{begin1}" then consume-linkage_consume else 0 end)/sum(case date when "{begin2}" then ad_click else 0 end),0),2) as "{begin3}"'''.format(begin1=i,begin2=i,begin3=i.replace("2020-",""))

        tmp3=''
        for i in daylist:  #消耗
            tmp3=tmp3+''' , round(ifnull(sum(case date when "{begin1}" then consume-linkage_consume else 0 end),0),0) as "{begin2}"'''.format(begin1=i,begin2=i.replace("2020-",""))
        tmpcvr = ''
        for i in daylist:  #转化率
            tmpcvr=tmpcvr+''' , concat(round(ifnull(sum(case date when "{begin1}" then ELT(b.effect_type, t1_num, t2_num,t3_num,t4_num,t5_num,t6_num,t7_num,t8_num,t9_num,t10_num,t11_num,t12_num,t13_num,t14_num,t15_num,t16_num) else 0 end)/sum(case date when "{begin2}" then ad_click else 0 end),0)*100,0),"%") as "{begin3}"'''.format(begin1=i,begin2=i,begin3=i.replace("2020-",""))
        tmp4=''' from tt.tt_advertiser_adzone aa
            left join  tt.tt_advertiser_effect_standard b on aa.advertiser_id=b.advertiser_id  left join (select (case lower(settle_type) when "cps" then "cps" else "媒购" end) as settle_type ,adzone_id from tt.tt_adzone_data group by adzone_id) tad on aa.adzone_id=tad.adzone_id 
            where date between "{begin}" and "{end}"'''.format(begin=self.begintime,end=self.endtime)
        if self.adzoneids<>'' and self.adzoneids<>'mytest':
            tmp5=''' and aa.adzone_id in ({0}) '''.format(self.adzoneids)
        else:
            tmp5=''
        tmp6='''group by aa.advertiser_id'''+tmpt03
        tmpsql=tmp+tmp1+tmp2+tmp3+tmpcvr+tmp4+tmp5+tmpt02+tmp6
        print tmpsql
        res,filed=db.selectsqlnew('devtidb',tmpsql)
        res = self.getresfloadtoint(res)
        self.exportexcel(filed,res,"reportmtpinggu")

        UnmatchedSQL = '''select ''' + UnmatchedHead + ''' ,"","","","","","","","",count(1),"","" '''+ UnmatchedEffectNum+ UnmatchedEffectNum1+ UnmatchedEffectNum1+ UnmatchedEffectNum1 +''' from voyagerlog.ad_effect_log_{month1} where (adzone_id is null or adzone_id="" or adzone_id=0 or advertiser_id is null or advertiser_id=""  or advertiser_id=0) and date_format(create_time,"%Y-%m-%d") between "{begin}" and "{end}" '''.format(month1=self.begintime[5:7],begin=self.begintime,end=self.endtime)

        ressum=""
        tmpsqlsum = '''(''' +UnmatchedSQL+''') union all('''+ tmpsql.replace(tmpt01,tmpt01sum).replace(tmpsql000,' ,"","","" ').replace(tmp6,"")+''')'''
        ressum, filedsum = db.selectsqlnew('devtidb', tmpsqlsum)
        ressum = self.getresfloadtoint(ressum)

        return res,filed,tmpsql+'''|||'''+tmpsqlsum,colspanx,ressum
        # return 1


    #菜单名：媒体效果-评估-小时表     #查询维度，日期
    def getmtpinguhour(self):

        tmpt01=''
        tmpt01sum=''
        tmpt02=''
        tmpt03=''
        tmpt04=''
        tmpt05=''
        tmpAd=''
        UnmatchedHead=''
        if self.advertiser_id<>'' and self.advertiser_id<>'mytest':
            tmpt01=''' azh.date as 日期,azh.advertiser_id as 广告主ID ,azh.advertiser_name as 广告主名称,azh.adzone_id as 广告位ID ,concat(azh.adzone_name,"&nbsp;-&nbsp;",ifnull(adz.settle_type,"")) as 广告位名称 '''
            tmpt01sum = ''' "","","","","汇总" '''
            UnmatchedHead=''' "","","","","未匹配效果数" '''
            tmpt02=''' and azh.advertiser_id in ({0}) '''.format(self.advertiser_id)
            tmpt03=''' ,azh.adzone_id '''
            tmpt04=''' left join (select adzone_id,sum(ifnull(platform_profit,0)) as platform_profit, (case lower(settle_type) when "cps" then "cps" else "媒购" end) as settle_type  from tt.tt_adzone_data where date between date_format(date_add(now() ,interval -7 day),'%Y-%m-%d') and  date_format(now(),'%Y-%m-%d') and adzone_cost>0 group by adzone_id) adz on adz.adzone_id=azh.adzone_id '''
            tmpt05=''' ,sum(ifnull(adz.platform_profit,0)) as platform_profit '''
            tmpAd=''' and a.advertiser_id in ({0}) '''.format(self.advertiser_id)
        else:
            tmpt01=''' azh.date as 日期,azh.advertiser_id as 广告主ID ,azh.advertiser_name as 广告主名称 '''
            tmpt01sum = ''' "","","汇总" '''
            UnmatchedHead=''' "","","未匹配效果数" '''
            tmpt05 = ''' ,999 as platform_profit '''

        if self.adzoneids <> '' and self.adzoneids <> 'mytest':
            tmpAz = ''' and adzone_id in ({0}) '''.format(self.adzoneids)
        else:
            tmpAz = ' '

        tmpsql000='''  ,ifnull(b.effect_typename,"") as 考核 ,ifnull(b.standard,"") as 要求 ,round(ifnull((sum(consume)-sum(linkage_consume))/sum(ELT(b.effect_type, t1_num, t2_num,t3_num,t4_num,t5_num,t6_num,t7_num,t8_num,t9_num,t10_num,t11_num,t12_num,t13_num,t14_num,t15_num,t16_num)),0),2) as 总成本,ifnull(tag,"") as 类型 '''
        tmpsql000sum='''  ,"","" ,round(ifnull((sum(consume)-sum(linkage_consume))/sum(ELT(b.effect_type, t1_num, t2_num,t3_num,t4_num,t5_num,t6_num,t7_num,t8_num,t9_num,t10_num,t11_num,t12_num,t13_num,t14_num,t15_num,t16_num)),0),2) as 总成本,"" '''

        tmpsql='''select  '''+ tmpt01 + tmpsql000 +'''
                ,round(ifnull(sum(ad_show),0),0) as 曝光
                ,round(ifnull(sum(ad_click),0),0) as 点击
                ,round(ifnull(sum(consume)-sum(linkage_consume),0),0) as 消耗
                ,round(ifnull((sum(consume)-sum(linkage_consume))/sum(ad_click),0),2) as CPC
                ,concat(round(ifnull(sum(ad_click)/sum(ad_show)*100,0),0),"%") as CTR
                ,round(ifnull((sum(consume)-sum(linkage_consume))/sum(ad_show)*1000,0),0) as CPM
                ,ifnull(sum(ELT(b.effect_type, t1_num, t2_num,t3_num,t4_num,t5_num,t6_num,t7_num,t8_num,t9_num,t10_num,t11_num,t12_num,t13_num,t14_num,t15_num,t16_num)),0) as 效果数
            ,concat(round(ifnull(sum(ELT(b.effect_type, t1_num, t2_num,t3_num,t4_num,t5_num,t6_num,t7_num,t8_num,t9_num,t10_num,t11_num,t12_num,t13_num,t14_num,t15_num,t16_num))/sum(ad_click),0)*100,2),"%") as 转化率
            '''
        tmpsql1=''
        tmpsqlCPM=''
        tmpsql2=''
        tmpsql3=''
        tmpsql6=''
        tmpcvr=''
        UnmatchedEffectNum=''
        UnmatchedEffectNum1=''

        for i in range(0,24):
            tmpsql1=tmpsql1+''' 
                             , round(ifnull(sum(case hour when {0} then consume-linkage_consume else 0 end)/
                                sum(case hour when {1} then
                                ELT(b.effect_type, t1_num, t2_num,t3_num,t4_num,t5_num,t6_num,t7_num,t8_num,t9_num,t10_num,t11_num,t12_num,t13_num,t14_num,t15_num,t16_num) 
                                else 0 end),0),2)
                              as "{2}"
                                '''.format(i,i,i)  #效果成本
            tmpsqlCPM=tmpsqlCPM+'''  , round(ifnull(sum(case hour when {0} then consume-linkage_consume else 0 end)/sum(case hour when {1} then ad_show else 0 end),0)*1000,0)  as "{2}" '''.format(i,i,i)  #CPM
            tmpsql2=tmpsql2+''' ,round(ifnull(sum(case hour when "{0}" then consume-linkage_consume else 0 end)/sum(case hour when "{1}" then ad_click else 0 end),0),2) as "{2}" '''.format(i,i,i)  #CPC
            tmpsql3=tmpsql3+''' ,round(ifnull(sum(case hour when "{0}" then consume-linkage_consume else 0 end),0),0) as "{1}" '''.format(i,i)  #消耗
            #转化率
            tmpcvr=tmpcvr+''' , concat(round(ifnull(sum(case hour when "{0}" then ELT(b.effect_type, t1_num, t2_num,t3_num,t4_num,t5_num,t6_num,t7_num,t8_num,t9_num,t10_num,t11_num,t12_num,t13_num,t14_num,t15_num,t16_num) else 0 end)/sum(case hour when "{1}" then ad_click else 0 end),0)*100,2),"%") as "{2}"'''.format(i,i,i)
            UnmatchedEffectNum = UnmatchedEffectNum + ''' ,sum(case hour(create_time) when {i} then 1 else 0 end) as "{i}" '''.format(i=i)
            UnmatchedEffectNum1 = UnmatchedEffectNum1 + ''' ,"" '''
        if self.begintime ==self.endtime and self.begintime==datetime.datetime.now().strftime('%Y-%m-%d')  :
            tmpTodayTable ='''(
                            select s.d    as date,
                                   s.h   as hour,
                                   s.advertiser_id as advertiser_id,
                                   (select name from voyager.advertiser where id = s.advertiser_id limit 1) as advertiser_name,
                                   (select (select name from voyager.industry where id = x.industry_id) from voyager.advertiser x where id = s.advertiser_id limit 1) as tag,
                                   s.adzone_id as adzone_id,
                                   (select adzone_name from voyager.base_adzone_info where id=s.adzone_id limit 1) as adzone_name,
                                   s.n as ad_show,
                                   c.n as ad_click,
                                   c.amount as consume,
                                   c.cashamount as cash_consume,
                                   c.linkageamount as linkage_consume,
                                   c.linkagecashamount as linkage_cash_consume,
                                   t1_num, t2_num, t3_num, t4_num, t5_num, t6_num, t7_num, t8_num, t9_num, t10_num, t11_num, t12_num, t13_num, t14_num, t15_num, t16_num
                            from (
                                     select date_format(create_time, '%Y-%m-%d') as d,hour(create_time) as h, advertiser_id, adzone_id, count(1) as n
                                     from ((select create_time,advertiser_id,adzone_id,status,position_id,"a" from voyagerlog.ad_show_log{day1}) union all(select create_time,advertiser_id,adzone_id,1,1,"b" from voyagerlog.ad_effect_log_{month0}  where date_format(create_time, '%Y-%m-%d')="{day11}")) a
                                     where status = 1
                                       and position_id <> 0 {ad1} {az1}
                                     group by date_format(create_time, '%Y-%m-%d'),hour(create_time), advertiser_id, adzone_id
                                 ) s left join (
                                     select date_format(a.create_time, '%Y-%m-%d') as d,hour(a.create_time) as h, a.advertiser_id, adzone_id, count(1) as n
                                          ,sum(charge_amount)/100 as amount
                                          ,sum(system_income + media_income_cash) / 100 as cashamount
                                          ,sum(case b.ocpa_ext_order when 1 then charge_amount  else 0 end) / 100    as linkageamount
                                          ,sum(case b.ocpa_ext_order when 1 then system_income + media_income_cash else 0 end) / 100   as linkagecashamount
                                     from voyagerlog.ad_click_log{day2} a left join voyager.ad_order b on a.ad_order_id=b.id
                                     where a.status = 1 {ad1} {az1}
                                     group by date_format(a.create_time, '%Y-%m-%d'),hour(a.create_time), a.advertiser_id, adzone_id
                                 ) c on s.d=c.d and s.h=c.h and s.advertiser_id=c.advertiser_id and s.adzone_id=c.adzone_id left join(
                                     select date_format(create_time, '%Y-%m-%d') as d,hour(create_time) as h, advertiser_id, adzone_id
                                          ,sum(case type when 1 then 1 else 0 end)  as t1_num
                                          ,sum(case type when 2 then 1 else 0 end)  as t2_num
                                          ,sum(case type when 3 then 1 else 0 end)  as t3_num
                                          ,sum(case type when 4 then 1 else 0 end)  as t4_num
                                          ,sum(case type when 5 then 1 else 0 end)  as t5_num
                                          ,sum(case type when 6 then 1 else 0 end)  as t6_num
                                          ,sum(case type when 7 then 1 else 0 end)  as t7_num
                                          ,sum(case type when 8 then 1 else 0 end)  as t8_num
                                          ,sum(case type when 9 then 1 else 0 end)  as t9_num
                                          ,sum(case type when 10 then 1 else 0 end)  as t10_num
                                          ,sum(case type when 11 then 1 else 0 end)  as t11_num
                                          ,sum(case type when 12 then 1 else 0 end)  as t12_num
                                          ,sum(case type when 13 then 1 else 0 end)  as t13_num
                                          ,sum(case type when 14 then 1 else 0 end)  as t14_num
                                          ,sum(case type when 15 then 1 else 0 end)  as t15_num
                                          ,sum(case type when 16 then 1 else 0 end)  as t16_num
                                     from voyagerlog.ad_effect_log_{month} a
                                     where  DATE_FORMAT(create_time, '%Y-%m-%d') = '{day3}'
                                        and ad_order_id not in (select id from voyager.ad_order where ocpa_ext_order = 1)
                                         {ad1} {az1}
                                     group by date_format(create_time, '%Y-%m-%d'),hour(create_time), advertiser_id, adzone_id
                                ) e on s.d=e.d and s.h=e.h and s.advertiser_id=e.advertiser_id and s.adzone_id=e.adzone_id 
                            )'''.format(day1=self.begintime.replace("-",""),month0=self.begintime[5:7],day11=self.begintime,day2=self.begintime.replace("-",""),month=self.begintime[5:7],day3=self.begintime,ad1=tmpAd,az1=tmpAz)
        else:
            tmpTodayTable = ''' tt.tt_advertiser_adzone_hour '''

        tmpsql4=''' from {0} azh
                left join  tt.tt_advertiser_effect_standard b on azh.advertiser_id=b.advertiser_id {1}  
                where ifnull(azh.consume,0)-ifnull(azh.linkage_consume,0) >0 and azh.date between "{2}" and "{3}" '''.format(tmpTodayTable,tmpt04,self.begintime,self.endtime)
        if self.adzoneids<>'' and self.adzoneids<>'mytest' :
            tmpsql5=''' and azh.adzone_id in ({0}) '''.format(self.adzoneids)
        else:
            tmpsql5=' '
        tmpsql6 = ''' group by azh.date,azh.advertiser_id '''+tmpt03
        tmpsqlall=tmpsql+tmpsql1+tmpsqlCPM+tmpsql2+tmpsql3+tmpcvr+tmpt05+tmpsql4+tmpsql5+tmpt02+tmpsql6

        res,filed=db.selectsqlnew('devtidb',tmpsqlall)
        res = self.getresfloadtoint(res)
        self.exportexcel(filed,res,"reportmtpingguhour")

        UnmatchedSQL = '''select ''' + UnmatchedHead + ''' ,"","","","","","","","","","",count(1),"" '''+ UnmatchedEffectNum+ UnmatchedEffectNum1+ UnmatchedEffectNum1+ UnmatchedEffectNum1+ UnmatchedEffectNum1 +''',999 from voyagerlog.ad_effect_log_{month1} where (adzone_id is null or adzone_id="" or adzone_id=0 or advertiser_id is null or advertiser_id=""  or advertiser_id=0) and date_format(create_time,"%Y-%m-%d") = "{begin}"  '''.format(month1=self.begintime[5:7],begin=self.begintime)

        ressum=""
        tmpsqlsum = '''(''' +UnmatchedSQL+''') union all('''+ tmpsqlall.replace(tmpt01,tmpt01sum).replace(tmpsql000,tmpsql000sum).replace(tmpsql6,"").replace(tmpt04,"").replace(tmpt05,",999")+''')'''

        ressum, filedsum = db.selectsqlnew('devtidb', tmpsqlsum)
        ressum = self.getresfloadtoint(ressum)

        return res,filed,tmpsqlall+'''|||'''+tmpsqlsum,ressum

    # 菜单名：媒体效果-评估-广告主类型
    def getreportByadvtag(self):
        daylist = self.gettimelist()

        if self.adzoneids <> '' and self.adzoneids <> 'mytest':
            tmpadzone = ''' and adzone_id in ({0}) '''.format(self.adzoneids)
        else:
            tmpadzone = ''


        tmpsql = '''select
                          aa.adzone_name as 广告位名称,aa.date as 日期,ifnull(tag, "")           as 类型
                         , round(ifnull(sum(ad_show), 0), 0)       as 曝光
                         , round(ifnull(sum(ad_click), 0), 0)      as 点击
                         , concat(round(ifnull(sum(ad_click)/sum(ad_show), 0)*100, 0),"%")      as CTR
                         , round(ifnull(sum(consume) - sum(linkage_consume), 0), 0)    as 消耗
                         , round(ifnull((sum(consume) - sum(linkage_consume)) / sum(ad_click), 0), 2)      as CPC
                         , ifnull(sum(ELT(b.effect_type, t1_num, t2_num, t3_num, t4_num, t5_num, t6_num, t7_num, t8_num, t9_num, t10_num,
                                          t11_num, t12_num, t13_num, t14_num, t15_num, t16_num)), 0)     as 效果数
                         , concat(round(ifnull(sum(ELT(b.effect_type, t1_num, t2_num, t3_num, t4_num, t5_num, t6_num, t7_num, t8_num,
                                                       t9_num, t10_num, t11_num, t12_num, t13_num, t14_num, t15_num, t16_num)) /
                                               sum(ad_click), 0) * 100, 0), "%")       as 转化率
                         , round(ifnull((sum(consume) - sum(linkage_consume)) / sum(
                            ELT(b.effect_type, t1_num, t2_num, t3_num, t4_num, t5_num, t6_num, t7_num, t8_num, t9_num, t10_num, t11_num,
                                t12_num, t13_num, t14_num, t15_num, t16_num)), 0), 2)      as 转化成本
                    from tt.tt_advertiser_adzone aa
                             left join tt.tt_advertiser_effect_standard b on aa.advertiser_id = b.advertiser_id
                    where date between "{begin}" and "{end}"
                         {adzone}
                    group by aa.tag,aa.date,aa.adzone_name
            '''.format(begin=self.begintime, end=self.endtime,adzone=tmpadzone)


        res, filed = db.selectsqlnew('devtidb', tmpsql)
        res = self.getresfloadtoint(res)
        self.exportexcel(filed, res, "reportByadvtag")

        return res, filed, tmpsql


    #菜单名：地域指标效果数据
    # 查询维度:日期，广告位，广告主
    def getregionbyday(self):
        daylist=self.gettimelist()
        tmpsql1='''select aar.advertiser_id as 广告主ID
                ,ifnull(aar.advertiser_name,"") as 广告主名称
                ,ifnull(b.effect_typename,"") as 考核
                ,ifnull(b.standard,"") as 要求
                ,ifnull(tag,"") as 类型
                ,ifnull(region,"") as 省
                ,round(ifnull(sum(ad_show),0),0) as 曝光
                ,round(ifnull((sum(consume)-sum(linkage_consume))/sum(ad_show)*1000,0),0) as CPM
                ,round(ifnull(sum(ad_click),0),0) as 点击
                ,round(ifnull(sum(consume)-sum(linkage_consume),0),0) as 消耗
                ,round(ifnull((sum(consume)-sum(linkage_consume))/sum(ad_click),0),2) as CPC
                ,concat(round(ifnull(sum(ad_click)/sum(ad_show),0)*100,0),"%") as CTR'''
        tmpsql2=''
        tmpsql3=''
        tmpsql4=''
        tmpsql5=''
        colspanx=0
        for i in daylist:
            colspanx=colspanx+1
            tmpsql2=tmpsql2+''' ,round(ifnull(sum(case date when "{0}" then consume-linkage_consume else 0 end)/
                         sum(case date when "{1}" then
                        ELT(b.effect_type, t1_num, t2_num,t3_num,t4_num,t5_num,t6_num,t7_num,t8_num,t9_num,t10_num,t11_num,t12_num,t13_num,t14_num,t15_num,t16_num)
                    else 0 end),0),2) as "{2}"'''.format(i,i,i.replace("2020-","")) #效果成本
            tmpsql3=tmpsql3+''' ,round(ifnull(sum(case date when "{0}" then consume-linkage_consume else 0 end),0),0) as "{1}"'''.format(i,i.replace("2020-","")) #消耗
            tmpsql4=tmpsql4+''' ,round(ifnull(sum(case date when "{0}" then consume-linkage_consume else 0 end)/sum(case date when "{1}" then ad_click else 0 end),0),2) as "{2}"'''.format(i,i,i.replace("2020-","")) #CPC
            tmpsql5=tmpsql5+''' ,round(ifnull(sum(case date when "{0}1" then adzone_click else 0 end),0),0) as "{1}"'''.format(i,i.replace("2020-","")) #入口点击
        tmpsql6=''' from tt_advertiser_adzone_region aar
                    left join  tt.tt_advertiser_effect_standard b on aar.advertiser_id=b.advertiser_id
                    where date between "{0}" and "{1}"'''.format(self.begintime,self.endtime)
        if self.adzoneids<>'' and self.adzoneids<>'mytest':
            tmpsql7=''' and adzone_id in ({0})'''.format(self.adzoneids)
        else:
            tmpsql7=' '
        # 必填
        tmpsql8=''' and aar.advertiser_id in ({0}) group by aar.advertiser_id,region;'''.format(self.advertiser_id)
        tmpall=tmpsql1+tmpsql2+tmpsql3+tmpsql4+tmpsql5+tmpsql6+tmpsql7+tmpsql8
        print tmpall
        res,filed=db.selectsqlnew('devtidb',tmpall)
        res = self.getresfloadtoint(res)
        self.exportexcel(filed,res,"reportregionbyday")
        return res,filed,tmpall,colspanx

    #菜单名：广告主地域效果
    #查询项：广告位ID，日期  #省（查出来在页面上）
    def getregionbyadv(self):
        daylist=self.gettimelist()
        tmp='''select aar.advertiser_id as 广告主ID
                ,ifnull(aar.advertiser_name,"") as 广告主名称
                ,ifnull(b.effect_typename,"") as 考核
                ,ifnull(b.standard,"") as 要求
                ,ifnull(tag,"") as 类型
                ,round(ifnull(sum(consume)-sum(linkage_consume),0),0) as 消耗
                ,round(ifnull((sum(consume)-sum(linkage_consume))/sum(ad_show)*1000,0),0) as CPM
                ,round(ifnull(sum(ad_click),0),0) as 点击
                ,round(ifnull((sum(consume)-sum(linkage_consume))/sum(ad_click),0),2) as CPC
                ,concat(round(ifnull(sum(
                        ELT(b.effect_type, t1_num, t2_num,t3_num,t4_num,t5_num,t6_num,t7_num,t8_num,t9_num,t10_num,t11_num,t12_num,t13_num,t14_num,t15_num,t16_num))
                        /sum( ad_click ),0)*100,0),"%") as CVR
                ,ifnull(region,"") as 省'''
        tmp1=''
        tmp2=''
        tmp3=''
        tmp4=''
        tmp5=''
        colspanx=0
        for i in daylist:
            colspanx=colspanx+1
            tmp1=tmp1+''' ,round(ifnull(sum(case date when "{0}" then consume-linkage_consume else 0 end)/
                 sum(case date when "{1}" then
                    ELT(b.effect_type, t1_num, t2_num,t3_num,t4_num,t5_num,t6_num,t7_num,t8_num,t9_num,t10_num,t11_num,t12_num,t13_num,t14_num,t15_num,t16_num)
                    else 0 end),0),2) as "{2}" '''.format(i,i,i.replace("2020-","")) #效果成本
            tmp2=tmp2+''' ,round(ifnull(sum(case date when "{0}" then consume-linkage_consume else 0 end),0),0) as "{1}"'''.format(i,i.replace("2020-","")) #消耗
            tmp3=tmp3+'''  ,round(ifnull(sum(case date when "{0}" then ad_click else 0 end),0),0) as "{1}"'''.format(i,i.replace("2020-","")) #点击
            tmp4=tmp4+''' ,round(ifnull(sum(case date when "{0}" then consume-linkage_consume else 0 end)/sum(case date when "{1}" then ad_click else 0 end),0),2) as "{2}"'''.format(i,i,i.replace("2020-","")) #CPC
            tmp5=tmp5+''',concat(round(ifnull(sum(case date when "{0}" then
                        ELT(b.effect_type, t1_num, t2_num,t3_num,t4_num,t5_num,t6_num,t7_num,t8_num,t9_num,t10_num,t11_num,t12_num,t13_num,t14_num,t15_num,t16_num)
                        else 0 end)
                        /sum(case date when "{1}" then ad_click else 0 end),0)*100,0),"%") as "{2}"'''.format(i,i,i.replace("2020-","")) #CVR
        tmp6='''from tt_advertiser_adzone_region aar
                left join  tt.tt_advertiser_effect_standard b on aar.advertiser_id=b.advertiser_id
                where date between "{0}" and "{1}"'''.format(self.begintime,self.endtime)
        if self.adzoneids<>'' and self.adzoneids<>'mytest':
            tmp7=''' and adzone_id in ({0})'''.format(self.adzoneids)
        else:
            tmp7=''
        if self.region<>'' and self.region<>'mytest':
            tmp8=''' and region like "%{0}%" '''.format(self.region)
        else:
            tmp8=''
        tmp9=' group by aar.advertiser_id,region;'
        tmpall=tmp+tmp1+tmp2+tmp3+tmp4+tmp5+tmp6+tmp7+tmp8+tmp9
        print tmpall
        res,filed=db.selectsqlnew('devtidb',tmpall)
        res = self.getresfloadtoint(res)
        self.exportexcel(filed,res,"reportregionbyadv")
        return res,filed,tmpall,colspanx



    # 菜单名：订单状态
    def getreportOrderState(self):

        if self.adzoneids <> '' :
            tmpadzone = ''' and adzone_id in ({0}) '''.format(self.adzoneids)
            #可投此位置的
            whereSQLadzone =''' 
                           and (adzone_id_direction = 0
                             or (adzone_id_direction = 1 and concat(",", adzone_id_content, ",") like "%,{0},%")
                             or (adzone_id_direction = 2 and concat(",", adzone_id_content, ",") not like "%,{1},%")
                             ) 
                             and (media_id_direction = 0
                                or (media_id_direction = 1 and concat(",", media_id_content, ",") like concat("%,",(select media_id from voyager.base_adzone_info where id={2} limit 1),",%"))
                                or (media_id_direction = 2 and concat(",", media_id_content, ",") not like concat("%,",(select media_id from voyager.base_adzone_info where id={3} limit 1),",%"))
                             ) 
                             '''.format(self.adzoneids,self.adzoneids,self.adzoneids,self.adzoneids)
        else:
            tmpadzone = ''
            whereSQLadzone = ''

        if self.isstatus == 'run':    #在投的
            whereSQL = ''' and state=4 '''+whereSQLadzone
        elif self.isstatus =='show':    #有曝光的
            whereSQL = '''  and id in (select ad_order_id from voyagerlog.ad_show_log{0}  where 1=1 {1}  ) '''.format(self.begintime.replace("-",""),tmpadzone)+whereSQLadzone
        else:
            whereSQL = ''


        tmpsql = '''select
                    ao.advertiser_id as 广告主ID,
                    concat((select name from voyager.advertiser where id=ao.advertiser_id limit 1),'<span class="redfont">',ao.托底,'</span>','<span class="redfont">',ao.ocpa_ext_order,'</span>','<span class="redfont">',ao.show_postion_fee,'</span>') as 广告主名称,
                    ifnull(asl.s,0) as 曝光,
                    ifnull(acl.c,0) as 点击,
                    ifnull(round(acl.amount,0),0) as 消耗,
                    ifnull(taen.num,0) as 效果数,
                    ifnull(concat(ifnull(round(taen.num/ro.c*100,0),0),"%"),0) as CVR,
                    ifnull(round(ro.consume/taen.num,2),0) as 实际成本,
                    ifnull(taes.standard,0) as 考核要求,
                    concat(ifnull(round(acl.c/asl.s*100,0),0),"%")  as CTR,
                    ifnull(round(acl.amount/asl.s*1000,0),0)  as CPM,
                    ifnull(round(acl.amount/acl.c,2),0)  as CPC,
                    ifnull(round(ao.payment,2),0)  as 出价,
                    ifnull(round(ro.ad_withhold,0),0) as 预扣,
                    ifnull(concat(round(acl.amount/ro.ad_withhold*100,0),"%"),0) as 消耗比例,
                    ao.state   as 状态,
                    ao.tag as   类型,
                    taes.effect_typename as 考核环节,
                    ao.run_type     as  竞价层级,
                    ao.id as 订单ID,
                    ao.name as 订单名称,
                    concat('<a href="',ac.image,'" target="_blank" >素材地址</a>') as 素材URL,
                    concat('<a href="',acel.link_common,'" target="_blank" >着落页</a>') as 着落页,
                    ac.审核状态,
                    ac.素材级别,
                    ao.push_order_admin    as  强推,
                    ao.advertise_start    as 投放开始日期,
                    ao.advertise_end  as 投放结束日期,
                    ao.advertise_time_type   as 投放时段类型,
                    concat('<input value="',replace(ao.advertise_time_type_detail,'"',''),'" />') as 投放时段,
                    ao.budget  as 每日预算（应该看预扣）,
                    ao.payment_mode as    投放类型,
                    ao.frequency  as    频次,
                    ao.budget_allocation   as   消耗速度,
                    ao.priority   as  权重,
                    ao.stop_time    as  订单暂停时间,
                    ao.quality_direction  as 媒体质量定向,
                    ao.media_id_direction    as  媒体定向,
                    concat('<input value="',ao.media_id_content,'" />')  as   媒体id,
                    ao.adzone_id_direction   as 广告位定向,
                    concat('<input value="',ao.adzone_id_content,'" />')      as  广告位id,
                    ao.adzone_limit  as  单位消耗上限,
                    ao.optimization_goal    as OCPA目标,
                    ao.target_cost  as  智能增量目标成本,
                    aod.region_direction   as 地域定向,
                    concat('<input value="',replace(aod.region,'"',''),'" />') as 地域,
                    aod.device_direction   as 设备定向,
                    aod.position_direction as 坑位定向
                from (
                         select id , advertiser_id , name
                              , (case (select id from voyager.config_parameters a where `desc` like '%托底%' and a.id = 20 and value like concat("%",aox.advertiser_id,"%")) when 20 then "托底" else "" end ) as 托底
                              , advertise_start    as advertise_start
                              , ELT(FIELD(advertise_long,0,1),advertise_end,"长期") as advertise_end
                              , ELT(advertise_time_type,"全时段","特定时段","高级")  as advertise_time_type
                              , ELT(advertise_time_type,"",concat(advertise_time_start,"至",advertise_time_end),advertise_time_perid) as advertise_time_type_detail
                              , budget / 100     as budget
                              , ELT(payment_mode,"CPM","CPC","CPA","OCPA") as payment_mode
                              , payment / 100  as payment
                              , ELT(FIELD(frequency_control,0,1),"无",frequency)  as  frequency
                              , ELT(budget_allocation,"标准","加速") as budget_allocation
                              , IFNULL(ELT(FIELD(state ,4,5,6),"投放中","暂停","结束"),"其他不可投放状态")  as state
                              , stop_time     as stop_time
                              , priority  as priority
                              , tag    as tag
                              , ELT(FIELD(quality_direction,0,1,2),"不限","普通","优媒")  as quality_direction
                              , ELT(FIELD(media_id_direction,0,1,2),"不限","媒体ID定向","媒体ID排除")   as media_id_direction
                              , media_id_content    as media_id_content
                              , ELT(FIELD(adzone_id_direction,0,1,2),"不限","广告位ID定向","广告位ID排除")   as adzone_id_direction
                              , adzone_id_content    as adzone_id_content
                              , adzone_limit   as adzone_limit
                              , ELT(run_type,"定向","通投")   as run_type
                              , ELT(optimization_goal,"表单预约","注册")    as optimization_goal
                              , ELT(FIELD(push_order_admin,1,0),"强推","")  as push_order_admin
                              , ELT(FIELD(ocpa_ext_order,1,0),"智能增量","")   as ocpa_ext_order
                              , case when ifnull(show_postion_fee,0) = 0 then "" else "只付费坑位" end as show_postion_fee
                              , target_cost    as target_cost
                         FROM voyager.ad_order aox
                         where 1=1
                         {whereSQL}
                 ) ao left join (
                    select order_id
                           ,ELT(FIELD(region_direction,0,1,2),"无","定向","排除")  as region_direction
                           ,region                            as region
                           ,ELT(FIELD(device_direction,0,1,2),"无","IOS","Android")   as device_direction
                           ,ifnull(ELT(FIELD(position_direction,0,1,2),"无","幸运奖弹层","天降红包"),position_direction)  as position_direction
                    from voyager.ad_order_direction
                    where 1=1 and is_valid = 1
                ) aod on ao.id=aod.order_id left join(
                    select order_id,creative_id  FROM voyager.ad_order_creative where state=1
                ) aoc on ao.id=aoc.order_id left join(
                    select creative_id,link_common  FROM voyager.ad_creative_link where is_valid=1
                ) acel on aoc.creative_id=acel.creative_id  left join (
                    select id,
                           image,
                           ELT(state, "待审核", "审核成功", "驳回") as 审核状态,
                           ELT(FIELD(level, 1, 3, 5, 10), "A1", "A3", "A5", "AX") as 素材级别,
                           ELT(creative_type, "图文", "视频") as 素材类型,
                           video_url
                    from voyager.ad_creative
                ) ac on aoc.creative_id=ac.id left join (
                    select count(1) as s,ad_order_id from voyagerlog.ad_show_log{begin1}  where status=1  {adzone_id1} group by ad_order_id
                ) asl on ao.id=asl.ad_order_id left join (
                    select count(1) as c,sum(charge_amount)/100 as amount,ad_order_id from voyagerlog.ad_click_log{begin2}  where status=1 and position_id<>0  {adzone_id2} group by ad_order_id
                ) acl on ao.id=acl.ad_order_id left join (
                    select adorder_id,
                           sum(show_num)           as s,
                           sum(adclick_num)        as c,
                           sum(ad_consume) / 100   as consume,
                           sum(cash_consume) / 100 as cash_consume,
                           sum(ad_withhold) / 100  as ad_withhold
                    from voyager.report_order
                    where  date = "{begin3}"
                    group by adorder_id
                ) ro on ao.id=ro.adorder_id left join (
                    select advertiser_id,effect_type,standard,effect_typename from tt.tt_advertiser_effect_standard
                ) taes on ao.advertiser_id=taes.advertiser_id left join (
                    select advertiser_id,ad_order_id,type,count(1) as num from voyagerlog.ad_effect_log_{month1} where date_format(create_time,"%Y-%m-%d")="{begin4}"  {adzone_id3}  group by ad_order_id,type
                ) taen on ao.id=taen.ad_order_id and taes.effect_type=taen.type 
            '''.format(whereSQL=whereSQL,begin1=self.begintime.replace("-",""),adzone_id1=tmpadzone,begin2=self.begintime.replace("-",""),adzone_id2=tmpadzone,begin3=self.begintime,month1=self.begintime[5:7],begin4=self.begintime,adzone_id3=tmpadzone)


        res, filed = db.selectsqlnew('devtidb', tmpsql)
        res = self.getresfloadtoint(res)
        #self.exportexcel(filed, res, "reportOrderState")

        return res, filed, tmpsql



    # 菜单名：广告位趋势
    def getreportZoneTrend(self):
        daylist=self.gettimelist()

        tmpsqlHead = '''
                select adzone_id as 广告位ID
                       ,(select adzone_name from voyager.base_adzone_info where id = rz.adzone_id limit 1)  as 广告位名称
                       ,ifnull(ELT(settle_method,"CPC","CPM","CPT","CPS","CPA","补量"),"") as 结算方式
                     '''
        tmpsqlHeadSum =''' select "","汇总","" '''
        colspanx = 0
        tmpsqlM1 = ""
        tmpsqlM2 = ""
        tmpsqlM3 = ""
        tmpsqlM4 = ""
        for i in daylist:
            colspanx=colspanx+1
            tmpsqlM1 = tmpsqlM1 + ''' ,round(ifnull(sum(case when date="{date}" then adzone_effect_num else 0 end),0),0) as "{dateTitle}" '''.format(date=i,dateTitle=i.replace("2020-",""))  #入口点击
            tmpsqlM2 = tmpsqlM2 + ''' ,round(ifnull(sum(case when date="{date}" then media_cost else 0 end)/100/sum(case when date="{date}" then adzone_effect_num else 0 end),0),3) as "{dateTitle}" '''.format(date=i,dateTitle=i.replace("2020-",""))   #入口成本
            tmpsqlM3 = tmpsqlM3 + ''' ,round(ifnull(sum(case when date="{date}" then platform_income + media_income_cash else 0 end)/100/sum(case when date="{date}" then adzone_effect_num else 0 end),0),3) as "{dateTitle}" '''.format(date=i,dateTitle=i.replace("2020-",""))  #入口收益
            tmpsqlM4 = tmpsqlM4 + ''' ,round(ifnull(sum(case when date="{date}" then adzone_consume else 0 end)/100,0),0) as "{dateTitle}" '''.format(date=i,dateTitle=i.replace("2020-",""))  #消耗
        tmpsqlFoot = '''
                from voyager.report_zone rz
                where date between "{begin}" and "{end}"
                group by adzone_id
            '''.format(begin=self.begintime,end=self.endtime)

        tmpsql = tmpsqlHead + tmpsqlM1 + tmpsqlM2 + tmpsqlM3 + tmpsqlM4 + tmpsqlFoot
        tmpsqlSum = tmpsqlHeadSum + tmpsqlM1 + tmpsqlM2 + tmpsqlM3 + tmpsqlM4 + tmpsqlFoot.replace("group by adzone_id","")

        res, filed = db.selectsqlnew('devtidb', tmpsql)
        res = self.getresfloadtoint(res)
        #self.exportexcel(filed, res, "reportZoneTrend")

        resSum, filedSum = db.selectsqlnew('devtidb', tmpsqlSum)
        resSum = self.getresfloadtoint(resSum)

        return res, filed, tmpsql , colspanx ,resSum





    # 菜单名：毛利预估
    def getreportPreProfitbyDay(self):
        daylist=self.gettimelist()

        tmpshowbaidu=''
        if self.showbaidu == '1':
            tmpshowbaidu = ''' and a.adzone_id in ({tmpBaiDuAdzoneID}) '''.format(tmpBaiDuAdzoneID=self.tmpBaiDuAdzoneID.replace(",,",""))
        else:
            tmpshowbaidu = ''

        adzonecost=""
        adzoneepc=""
        adzonePassIF=""
        adzoneDetail=""
        for i in daylist:
            adzonecost = adzonecost + ''' ,{dateTitle}adzonecost '''.format(dateTitle=i.replace("2020-","").replace("-",""))
            adzoneepc= adzoneepc + ''' ,{dateTitle}adzoneepc '''.format(dateTitle=i.replace("2020-","").replace("-",""))
        for i in daylist:
            adzonePassIF = adzonePassIF + ''' when ifnull(ELT(datediff(a.date,"{date}") {adzonecost}),0)>0
                                         then ELT(datediff(a.date,"{date}") {adzoneepc})*a.ad_show
                                          '''.format(date=i,adzonecost=adzonecost,adzoneepc=adzoneepc)
            adzoneDetail = adzoneDetail + ''' , sum(case when date="{date}" then adzone_click else 0 end) as {dateTitle}adzoneclick
                                      , sum(case when date="{date}" then ad_show else 0 end) as {dateTitle}adshow
                                      , sum(case when date="{date}" and adzone_click>100 then adzone_cost else 0 end) as {dateTitle}adzonecost
                                      , sum(case when date="{date}" then adzone_cost else 0 end)/sum(case when date="{date}" then adzone_click else 0 end) as {dateTitle}adzoneepc 
                                      '''.format(date=i,dateTitle=i.replace("2020-","").replace("-",""))

        tmpsql = '''
                 select a.date
                     ,round(ifnull(sum(a.ad_show),0),0) as 广告曝光
                     ,round(ifnull(sum(a.cash_consume),0),0) as 现金消耗
                     ,round(ifnull(sum(a.adzone_cost),0),0) as 媒体成本
                     ,round(ifnull(sum(a.pre_adzone_cost),0),0) as 预估媒体成本
                     ,round(ifnull(sum(platform_profit),0),0) as 平台毛利
                     ,round(ifnull(sum(a.cash_consume)-sum(a.pre_adzone_cost),0),0) as 预估平台毛利
                     ,round(ifnull(sum(case when a.pre_adzone_cost=0 then  a.cash_consume-a.pre_adzone_cost else 0 end),0),0) as 无预估媒体成本的消耗即毛利值
                 from (
                
                             select a.date as date,a.adzone_id
                                 ,sum(a.ad_show) as ad_show
                                 ,sum(a.cash_consume) as cash_consume
                                 ,sum(a.adzone_cost) as adzone_cost
                                 ,sum(platform_profit) as platform_profit
                                 ,sum(case when ifnull((a.adzone_cost),0)<>0  then (a.adzone_cost)  
                                     {adzonePassIF}
                                     else ifnull((a.adzone_cost),0)
                                 end) as pre_adzone_cost
                             from tt.tt_adzone_data a
                                left join (
                                     select  adzone_id {adzoneDetail}
                                     from tt.tt_adzone_data
                                     where date between "{beginPass}" and "{end}"
                                     group by adzone_id
                                ) b on  a.adzone_id=b.adzone_id
                             where a.date between "{begin}" and "{end}"
                             group by a.date,a.adzone_id
                
                          ) a
                 where a.date between "{begin}" and "{end}"
                 {tmpshowbaidu}
                 group by a.date
                 order by a.date
            '''.format(beginPass=self.begintime,begin=self.begintime,end=self.endtime,adzonePassIF=adzonePassIF,adzoneDetail=adzoneDetail,tmpshowbaidu=tmpshowbaidu)


        res, filed = db.selectsqlnew('devtidb', tmpsql)
        res = self.getresfloadtoint(res)
        self.exportexcel(filed, res, "reportPreProfitbyDay")

        return res, filed, tmpsql




        # 菜单名：活动维度广告主效果
    def getreportAdzoneActEffect(self):
        daylist = self.gettimelist()

        tmpad = ''
        tmpadz = ''
        tmpact = ''
        if self.advertiser_id <> '' and self.advertiser_id <> 'mytest':
            tmpad = ''' and a.advertiser_id in ({0}) '''.format(self.advertiser_id)
        else:
            tmpad = ' '

        if self.adzoneids <> '' and self.adzoneids <> 'mytest':
            tmpadz = ''' and a.adzone_id in ({0}) '''.format(self.adzoneids)
        else:
            tmpadz = ''

        if self.actid <> '' and self.actid <> 'mytest':
            tmpact = ''' and a.act_id in ({0}) '''.format(self.actid)
        else:
            tmpact = ''

        colspanx = 0
        tmpIIR = ''
        tmpADShow = ''
        tmpADClick = ''
        tmpCPC = ''
        tmpConsume = ''
        tmpEffectCost = ''
        tmpEffectNum = ''
        tmpActShow = ''
        for i in daylist:  # 转化成本
            colspanx = colspanx + 1
            tmpIIR = tmpIIR + ''' , round(ifnull(sum(case when date="{begin1}" then a.iir else 0 end),0),2) as "{begin2}" '''.format(begin1=i, begin2=i.replace("2020-", "")) #IIR
            tmpADShow = tmpADShow + ''' , ifnull(sum(case when date="{begin1}" then a.ad_show else 0 end),0) as "{begin2}" '''.format(begin1=i, begin2=i.replace("2020-", "")) #ADShow
            tmpADClick = tmpADClick + ''' , ifnull(sum(case when date="{begin1}" then a.ad_click else 0 end),0) as "{begin2}" '''.format(begin1=i, begin2=i.replace("2020-", "")) #ADClick
            tmpCPC = tmpCPC + ''' , round(ifnull(sum(case when date="{begin1}" then a.nolinkconsume else 0 end),0)/ifnull(sum(case when date="{begin1}" then a.ad_click else 0 end),0),2) as "{begin2}" '''.format(begin1=i, begin2=i.replace("2020-", "")) #CPC
            tmpConsume = tmpConsume + ''' , round(ifnull(sum(case when date="{begin1}" then a.nolinkconsume else 0 end),0),0) as "{begin2}" '''.format(begin1=i, begin2=i.replace("2020-", "")) #Consume
            tmpEffectCost = tmpEffectCost + ''' , round(ifnull(sum(case when date="{begin1}" then a.nolinkconsume else 0 end),0)/ifnull(sum(case when date="{begin1}" then a.enum else 0 end),0),2) as "{begin2}" '''.format(begin1=i, begin2=i.replace("2020-", "")) #EffectCost
            tmpEffectNum = tmpEffectNum + ''' , ifnull(sum(case when date="{begin1}" then a.enum else 0 end),0) as "{begin2}" '''.format(begin1=i, begin2=i.replace("2020-", "")) #EffectNum
            if colspanx==1:
                tmpActShow = tmpActShow + ''' (select date_format(create_time,"%Y-%m-%d") as date,adzone_id,act_id,count(1) as s from voyagerlog.act_click_log{begin} where status=1 group by adzone_id,act_id)  '''.format(begin=i.replace("-", "")) #tmpActShow
            else:
                tmpActShow = tmpActShow + ''' union all (select date_format(create_time,"%Y-%m-%d") as date,adzone_id,act_id,count(1) as s from voyagerlog.act_click_log{begin} where status=1 group by adzone_id,act_id) '''.format(begin=i.replace("-", "")) #tmpActShow

        tmpHead = ''' select   concat(a.act_name,"(",a.act_id,")") as 活动名称
                              , a.advertiser_id as 广告主ID
                              , a.advertiser_name as 广告主名称
                              , a.tag as 广告类别
                              , concat(a.adzone_name,"(",a.adzone_id,")") as 广告位名称
                 '''
        tmp = '''            {tmpIIR} {tmpADShow} {tmpADClick} {tmpCPC} {tmpConsume} {tmpEffectCost} {tmpEffectNum}
                from (
                         select a.date
                              , a.advertiser_id
                              , a.advertiser_name
                              , a.tag
                              , a.adzone_id
                              , a.adzone_name
                              , a.act_id
                              , a.act_name
                              , ifnull(d.ad_show / c.s,0)   as IIR
                              , ifnull(sum(a.ad_show),0)  as ad_show
                              , ifnull(sum(a.ad_click),0)  as ad_click
                              , ifnull(sum(a.consume - a.linkage_consume),0)   as nolinkconsume
                              , ifnull(sum(ELT(b.effect_type, t1_num, t2_num, t3_num, t4_num, t5_num, t6_num, t7_num, t8_num, t9_num, t10_num, t11_num, t12_num, t13_num, t14_num, t15_num, t16_num)),0) as enum
                         from tt.tt_advertiser_act a
                                  left join tt.tt_advertiser_effect_standard b on a.advertiser_id = b.advertiser_id
                                  left join 
                                   ({tmpActShow}) c on a.adzone_id = c.adzone_id and a.date = c.date  and a.act_id=c.act_id
                                   left join (
                                     select a.date , a.advertiser_id , a.adzone_id , a.act_id , ifnull(sum(a.ad_show),0)  as ad_show
                                     from tt.tt_advertiser_act a
                                     where a.date between "{begin1}" and "{end1}"
                                     group by a.date,a.adzone_id,a.act_id
                                    ) d on a.adzone_id = d.adzone_id and a.date = d.date  and a.act_id=d.act_id
                         where a.date between "{begin}" and "{end}"
                         group by a.date,
                                  a.advertiser_id,
                                  a.adzone_id,
                                  a.act_id
                     ) a 
                     where 1=1 {tmpad} {tmpadz} {tmpact}
                 '''.format(begin=self.begintime, end=self.endtime,begin1=self.begintime, end1=self.endtime,tmpIIR=tmpIIR,tmpADShow=tmpADShow,tmpADClick=tmpADClick,tmpCPC=tmpCPC,tmpConsume=tmpConsume,tmpEffectCost=tmpEffectCost,tmpEffectNum=tmpEffectNum,tmpad=tmpad,tmpadz=tmpadz,tmpact=tmpact,tmpActShow=tmpActShow)
        tmpFood = ''' group by  a.advertiser_id, a.adzone_id, a.act_id  '''
        tmpsql = tmpHead + tmp + tmpFood

        res, filed = db.selectsqlnew('devtidb', tmpsql)
        res = self.getresfloadtoint(res)
        self.exportexcel(filed, res, "reportAdzoneActEffect")

        tmpsqlsum = ''' select "","","","","" ''' + tmp
        ressum, filedsum = db.selectsqlnew('devtidb', tmpsqlsum)
        ressum = self.getresfloadtoint(ressum)

        return res, filed, tmpsql, colspanx, ressum

    #易积分后台报表    菜单名---- 易积分后台报表
    def getyijifen(self):
        headtr=''
        # 全部报表
        if self.type=='0':
            # yjf_dau	'媒体DAU',' '  商城pv,' ' 商城uv,total_amount	'总订单数',total_income	'总收益',arpu	'arpu',valid_amount	'有效订单数',valid_payment	'有效实付金额',valid_cost	'有效成本',valid_income	'有效收益',settle_amount	'结算订单数',settle_payment	'结算实付金额',settle_cost	'结算成本',settle_income	'结算收益',update_time,
            tmpsql='''SELECT app_id	'应用id',app_name '应用名称',riqi	'日期',
                    yjf_dau	'媒体DAU',' '  商城pv,' ' 商城uv,total_income	'总收益',arpu	'arpu',update_time,
                    yjf_earn_yjf_val	'赚易积分数量',yjf_consume_yjf_val	'消易积分数量',yjf_consume_yjf_uv	'消易积分人数',' ' 积分价值,
                    hdt_media_id	'媒体id',hdt_income	'广告收益',hdt_voyager_income	'互动推收益',hdt_gdt_income	'广点通收益',hdt_csj_income	'穿山甲收益',hdt_adzone_uv	'广告位uv',hdt_lottery_uv	'抽奖uv',hdt_adshow_uv	'广告曝光uv',hdt_adclick_uv	'广告点击uv',hdt_adzone_effect_num	'广告位有效点击数量',hdt_lottery_effect_num	'抽奖有效点击数量',hdt_ad_show_effect_num	'广告曝光有效点击数量',hdt_ad_click_effect_num	'广告点击有效点击数量',
                    shqy_total_amount	'总订单数',shqy_total_payment	'总实付金额',shqy_total_cost	'总成本',shqy_total_income	'总收益',shqy_valid_amount	'有效订单数',shqy_valid_payment	'有效实付金额',shqy_valid_cost	'有效成本',shqy_valid_income	'有效收益',shqy_settle_amount	'结算订单数',shqy_settle_payment	'结算实付金额',shqy_settle_cost	'结算成本',shqy_settle_income	'结算收益',
                    yhxb_total_amount	'总订单数',yhxb_total_payment	'总实付金额',yhxb_total_cost	'总成本',yhxb_total_income	'总收益',yhxb_valid_amount	'有效订单数',yhxb_valid_payment	'有效实付金额',yhxb_valid_cost	'有效成本',yhxb_valid_income	'有效收益',yhxb_settle_amount	'结算订单数',yhxb_settle_payment	'结算实付金额',yhxb_settle_cost	'结算成本',yhxb_settle_income	'结算收益',
                    ygbb_total_amount	'总订单数',ygbb_total_payment	'总实付金额',ygbb_total_cost	'总成本',ygbb_total_income	'总收益',ygbb_valid_amount	'有效订单数',ygbb_valid_payment	'有效实付金额',ygbb_valid_cost	'有效成本',ygbb_valid_income	'有效收益',ygbb_settle_amount	'结算订单数',ygbb_settle_payment	'结算实付金额',ygbb_settle_cost	'结算成本',ygbb_settle_income	'结算收益'
                    from yjf_data_report  where riqi>='{begin1}' and riqi<='{end1}';
            '''.format(begin1=self.begintime,end1=self.endtime)
            headtr='''<tr>
        <td>&nbsp;</td>
        <td>&nbsp;</td>
        <td>&nbsp;</td>
        <td align="center" colspan=6>汇总</td>
        <td align="center" colspan=4>积分运营</td>
        <td align="center" colspan=14>互动推</td>
        <td align="center" colspan=12>生活权益</td>
        <td align="center" colspan=12>优惠线报</td>
        <td align="center" colspan=12>易购宝贝</td></tr>
            '''
        #     生活权益
        elif self.type=='1':
            tmpsql='''SELECT app_id 应用id,riqi 日期, total_amount 总订单数,ROUND(total_payment/1,2) 总实付金额,
                    ROUND(total_cost/1,2) 总成本, ROUND(total_income/1,2) 总收益,valid_amount 有效订单数,ROUND(valid_payment/1,2) 有效实付金额,ROUND(valid_cost/1,2) 有效成本,valid_income 有效收益 ,
                     settle_amount 结算订单数,ROUND(settle_payment/1,2) 结算实付金额,ROUND(settle_cost/1,2) 结算成本,settle_income 结算收益,update_time
                        from yjf_origin_order_pool_shqy where riqi>='{begin1}' and riqi<='{end1}'
            '''.format(begin1=self.begintime,end1=self.endtime)
        #     优惠线报
        elif self.type=='2':
            tmpsql='''SELECT app_id 应用id,riqi 日期, total_amount 总订单数,ROUND(total_payment/1,2) 总实付金额,
                    ROUND(total_cost/1,2) 总成本, ROUND(total_income/1,2) 总收益,valid_amount 有效订单数,ROUND(valid_payment/1,2) 有效实付金额,ROUND(valid_cost/1,2) 有效成本,valid_income 有效收益 ,
                     settle_amount 结算订单数,ROUND(settle_payment/1,2) 结算实付金额,ROUND(settle_cost/1,2) 结算成本,settle_income 结算收益,update_time
                        from yjf_origin_order_yhxb where riqi>='{begin1}' and riqi<='{end1}'
            '''.format(begin1=self.begintime,end1=self.endtime)
        #     易购宝贝
        elif self.type=='3':
            tmpsql='''SELECT app_id 应用id,riqi 日期, total_amount 总订单数,ROUND(total_payment/1,2) 总实付金额,
                    ROUND(total_cost/1,2) 总成本, ROUND(total_income/1,2) 总收益,valid_amount 有效订单数,ROUND(valid_payment/1,2) 有效实付金额,ROUND(valid_cost/1,2) 有效成本,valid_income 有效收益 ,
                     settle_amount 结算订单数,ROUND(settle_payment/1,2) 结算实付金额,ROUND(settle_cost/1,2) 结算成本,settle_income 结算收益,update_time
                        from yjf_origin_order_ygbb where riqi>='{begin1}' and riqi<='{end1}'
            '''.format(begin1=self.begintime,end1=self.endtime)
        #     互动推广告
        elif self.type=='4':
            tmpsql='''SELECT app_id 应用id,riqi 日期,media_id 媒体id,ROUND(income/1,2) 广告收益,ROUND(voyager_income/1,2) 互动推收益,ROUND(gdt_income/1,2) 广点通收益,ROUND(csj_income/1,2) 穿山甲收益,adzone_uv 广告位uv,lottery_uv 抽奖uv,
                    adshow_uv 广告曝光uv,adclick_uv 广告点击uv,adzone_effect_num 广告位有效点击数量,lottery_effect_num 抽奖有效点击数量,ad_show_effect_num 广告曝光有效点击数量,ad_click_effect_num 广告点击有效点击数量
                     from yjf_origin_report_voyager where riqi>='{begin1}' and riqi<='{end1}'
            '''.format(begin1=self.begintime,end1=self.endtime)
        #     易积分
        elif self.type=='5':
            tmpsql='''SELECT app_id 应用id,riqi 日期,dau 日活,earn_yjf_val 赚易积分数量,consume_yjf_uv 消易积分人数 ,update_time
                        from yjf_origin_report_yjf  where riqi>='{begin1}' and riqi<='{end1}'
            '''.format(begin1=self.begintime,end1=self.endtime)
        tmpdit={"媒体DAU":"媒体日活-统计当日该应用下设备号数量（需去重）"}
        res,filed=db.selectsqlnew('devtidb',tmpsql)
        res=self.getresfloadtoint(res)
        self.exportexcel(filed,res,"reportgetyijifen")
        filed=self.getfieldname(filed,tmpdit)
        return res,filed,tmpsql,headtr

    #易积分后台报表    菜单名---- 易积分后台报表
    def getyijifenall(self):
        headtr=''
        # 全部报表
        if self.type=='0':
            # yjf_dau	'媒体DAU',' '  商城pv,' ' 商城uv,total_amount	'总订单数',total_income	'总收益',arpu	'arpu',valid_amount	'有效订单数',valid_payment	'有效实付金额',valid_cost	'有效成本',valid_income	'有效收益',settle_amount	'结算订单数',settle_payment	'结算实付金额',settle_cost	'结算成本',settle_income	'结算收益',update_time,
            tmpsql='''SELECT yjf_data_report_id,riqi '日期',app_id '应用id',app_name '应用名称',
                    concat('<input type="number" oninput="value=value.replace(/[^\d]/g,'')" id="row-3-age" name="row-4-age" onblur="saveTabeleData(this',',\\'pv\\'',')" value="',ifnull(yjf_shangcheng_pv,''),'"/>') '商城pv',
                    concat('<input type="number" oninput="value=value.replace(/[^\d]/g,'')"  id="row-4-age" name="row-4-age" onblur="saveTabeleData(this',',\\'uv\\'',')" value="',ifnull(yjf_shangcheng_uv,''),'"/>') '商城uv',
                    total_income '总收益',
                    CASE  WHEN yjf_shangcheng_uv=0 THEN 0 WHEN yjf_shangcheng_uv>=0 THEN ROUND(total_income/yjf_shangcheng_uv,2) END 'arpu',
                    hdt_adzone_uv '广告位uv',
                    CASE WHEN yjf_shangcheng_uv=0 THEN 0 WHEN yjf_shangcheng_uv>=0 THEN   ROUND(hdt_adzone_uv/yjf_shangcheng_uv,2) END '广告位点击率',
                    hdt_lottery_uv '抽奖uv',
                    CASE WHEN hdt_adzone_uv=0 THEN 0 WHEN hdt_adzone_uv>=0 THEN ROUND(hdt_lottery_uv/hdt_adzone_uv,2) END  '参与率',
                    hdt_adshow_uv '广告曝光uv',hdt_ad_show_effect_num '广告曝光数',
                    CASE WHEN hdt_adshow_uv=0 THEN 0 WHEN hdt_adshow_uv>=0 THEN ROUND(hdt_ad_show_effect_num/hdt_adshow_uv,2) END  '人均广告曝光次数',
                    hdt_adclick_uv '广告点击uv',hdt_ad_click_effect_num '广告点击数',
                    CASE WHEN hdt_ad_show_effect_num=0 THEN 0 WHEN hdt_ad_show_effect_num>=0 THEN ROUND(hdt_ad_click_effect_num/hdt_ad_show_effect_num,2) END  '点击率CTR' ,
                    hdt_income '广告收益',hdt_voyager_income '互动推收益',hdt_voyager_adshow '互动推广告曝光',
                    CASE WHEN hdt_voyager_adshow=0 THEN 0 WHEN hdt_voyager_adshow>=0 THEN ROUND(hdt_voyager_income/(hdt_voyager_adshow*1000),2) END  '互动推CPM' ,
                    hdt_gdt_income '广点通收益',hdt_gdt_adshow '广点通曝光',
                    CASE WHEN hdt_gdt_adshow=0 THEN 0 WHEN hdt_gdt_adshow>=0 THEN ROUND(hdt_gdt_income/(hdt_gdt_adshow*1000),2) END  '广点通CPM' ,
					hdt_csj_income '穿山甲收益', hdt_csj_adshow '穿山甲曝光',
                    CASE WHEN hdt_csj_adshow=0 THEN 0 WHEN hdt_csj_adshow>=0 THEN ROUND(hdt_csj_income/(hdt_csj_adshow*1000),2) END  '穿山甲CPM' ,
					valid_amount '有效订单数',yjf_earn_yjf_val '赚积分金额',yjf_consume_yjf_val '消积分金额'
                    from tt.yjf_data_report
                    where app_name<>''and riqi>='{begin1}' and riqi<='{end1}';
                     '''.format(begin1=self.begintime,end1=self.endtime)

        #     生活权益
        elif self.type=='1':
            tmpsql='''SELECT app_id 应用id,riqi 日期, total_amount 总订单数,ROUND(total_payment/1,2) 总实付金额,
                    ROUND(total_cost/1,2) 总成本, ROUND(total_income/1,2) 总收益,valid_amount 有效订单数,ROUND(valid_payment/1,2) 有效实付金额,ROUND(valid_cost/1,2) 有效成本,valid_income 有效收益 ,
                     settle_amount 结算订单数,ROUND(settle_payment/1,2) 结算实付金额,ROUND(settle_cost/1,2) 结算成本,settle_income 结算收益,update_time
                        from yjf_origin_order_pool_shqy where riqi>='{begin1}' and riqi<='{end1}'
            '''.format(begin1=self.begintime,end1=self.endtime)
        #     优惠线报
        elif self.type=='2':
            tmpsql='''SELECT app_id 应用id,riqi 日期, total_amount 总订单数,ROUND(total_payment/1,2) 总实付金额,
                    ROUND(total_cost/1,2) 总成本, ROUND(total_income/1,2) 总收益,valid_amount 有效订单数,ROUND(valid_payment/1,2) 有效实付金额,ROUND(valid_cost/1,2) 有效成本,valid_income 有效收益 ,
                     settle_amount 结算订单数,ROUND(settle_payment/1,2) 结算实付金额,ROUND(settle_cost/1,2) 结算成本,settle_income 结算收益,update_time
                        from yjf_origin_order_yhxb where riqi>='{begin1}' and riqi<='{end1}'
            '''.format(begin1=self.begintime,end1=self.endtime)
        #     易购宝贝
        elif self.type=='3':
            tmpsql='''SELECT app_id 应用id,riqi 日期, total_amount 总订单数,ROUND(total_payment/1,2) 总实付金额,
                    ROUND(total_cost/1,2) 总成本, ROUND(total_income/1,2) 总收益,valid_amount 有效订单数,ROUND(valid_payment/1,2) 有效实付金额,ROUND(valid_cost/1,2) 有效成本,valid_income 有效收益 ,
                     settle_amount 结算订单数,ROUND(settle_payment/1,2) 结算实付金额,ROUND(settle_cost/1,2) 结算成本,settle_income 结算收益,update_time
                        from yjf_origin_order_ygbb where riqi>='{begin1}' and riqi<='{end1}'
            '''.format(begin1=self.begintime,end1=self.endtime)
        #     互动推广告
        elif self.type=='4':
            tmpsql='''SELECT app_id 应用id,riqi 日期,media_id 媒体id,ROUND(income/1,2) 广告收益,ROUND(voyager_income/1,2) 互动推收益,ROUND(gdt_income/1,2) 广点通收益,ROUND(csj_income/1,2) 穿山甲收益,adzone_uv 广告位uv,lottery_uv 抽奖uv,
                    adshow_uv 广告曝光uv,adclick_uv 广告点击uv,adzone_effect_num 广告位有效点击数量,lottery_effect_num 抽奖有效点击数量,ad_show_effect_num 广告曝光有效点击数量,ad_click_effect_num 广告点击有效点击数量
                     from yjf_origin_report_voyager where riqi>='{begin1}' and riqi<='{end1}'
            '''.format(begin1=self.begintime,end1=self.endtime)
        #     易积分
        elif self.type=='5':
            tmpsql='''SELECT app_id 应用id,riqi 日期,dau 日活,earn_yjf_val 赚易积分数量,consume_yjf_uv 消易积分人数 ,update_time
                        from yjf_origin_report_yjf  where riqi>='{begin1}' and riqi<='{end1}'
            '''.format(begin1=self.begintime,end1=self.endtime)
        tmpdit={"媒体DAU":"媒体日活-统计当日该应用下设备号数量（需去重）"}
        current_app.logger.warning("{0}".format(tmpsql))
        res,filed=db.selectsqlnew('devtidb',tmpsql)
        res=self.getresfloadtoint(res)
        self.exportexcel(filed,res,"reportgetyijifen")
        filed=self.getfieldname(filed,tmpdit)
        return res,filed,tmpsql,headtr
    def yjf_update(self):
        if self.update_type=='shangchengpv':
            tmpsql='''
            UPDATE yjf_data_report set yjf_shangcheng_pv={yjf_shangcheng_pv} where yjf_data_report_id={yjf_data_report_id}
            '''.format(yjf_shangcheng_pv=self.value,yjf_data_report_id=self.yjf_data_report_id)

            # tmpsql='''
            # UPDATE yjf_app_inter_config set app_name='{yjf_shangcheng_pv}' where yjf_data_report_id={yjf_data_report_id}
            # '''.format(yjf_shangcheng_pv=self.value,yjf_data_report_id=self.yjf_data_report_id)
        else:
            tmpsql='''
            UPDATE yjf_data_report set yjf_shangcheng_uv={yjf_shangcheng_uv} where yjf_data_report_id={yjf_data_report_id}
            '''.format(yjf_shangcheng_uv=self.value,yjf_data_report_id=self.yjf_data_report_id)
        try:
            db.execsql('devtidb',tmpsql)
        except Exception as e:
            current_app.logger.warning("update yjf_app_inter_config fail error is  {0}".format(e.message))
        return 1






if __name__ == '__main__':
    # test=myreport(begintime='2020-04-1',endtime='2020-04-02',adzoneids='21')
    # test=myreport(begintime='2020-04-01',endtime='2020-04-11',adzoneids='22222',advertiser_id=1,region='北京')
    # tmp=test.getallreport()
    tmp='''<td title="{0}">状态</font></td>'''
    fild=['a','b','c']
    tmpdit={'a':'test'}
    tmplist=[]

    fild=[tmp.format(tmpdit[x]) if x in tmpdit.keys() else x for x in fild]
    # print fild
    #
    # for i in fild:
    #     if i in tmpdit.keys():
    #         i=tmp.format(tmpdit[i])
    #         tmplist.append(i)
    #     else:
    #         tmplist.append(i)
    print fild


    # print tmp
    # print get_date_list('2018-01-01','2018-02-28')
    # cwd = os.getcwd()
    # print(cwd)
    # wb=Workbook()
    # sheet=wb.active
    # wb1=load_workbook('reportall.xlsx')
    # w=wb1["Sheet"]
    # sheet.merge_cells('J1:L1')
    # sheet['J1']='消费消费'



