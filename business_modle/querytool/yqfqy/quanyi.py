# -*- coding: utf-8 -*-
__author__ = 'jinyi'
from business_modle.querytool import db
import requests as r
import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import  Workbook,load_workbook
import os

class quanyi(object):
    def __init__(self,env=''):
        self.env=env
    # 导出excel
    def exportexcel(self,filed,res,excelname):
        if len(res)<>0:
            res=list(res)
            # res.insert(0,headtr)
            res.insert(0,filed)
            wb=Workbook()
            sheet=wb.active
            for i in range(len(res)):
                sheet.append(res[i])
            try:
                # wb.save("../../../static/result/reportall.xlsx")
                wb.save("./static/result/{0}.xlsx".format(excelname))
            except Exception as e:
                print e.message
        return ''
    def dubtolist(self,re):
        tmp=[]
        for i in re:
            tmp.append(list(i))
        return tmp
    def selectre(self,tmpsql):
        if self.env=='test':
            res,filed=db.selectsqlnew('testquanyi',tmpsql)
            res=self.dubtolist(res)
        else:
            res,filed=db.selectsqlnew('devquanyi',tmpsql)
            res=self.dubtolist(res)
        return  res,filed
    # 查询权益订单
    def georder(self):
        tmpsql='''
               SELECT o.ORDER_NO 订单对外唯一标识,o.OPEN_ID 微信openid,o.PRODUCT_ID 商品id ,o.SHOP_SHORT_NAME 商家名称 ,o.SHOP_PRODUCT_ID 商家商品id ,
                o.RECOMMEND_WEBSITE_ID 推荐人站点id ,o.FEEDBACK_TAG 反馈标签 ,o.PRICE_CURRENT 商品现价 ,o.TOTAL_AMOUNT 购买商品数量 ,
                o.TOTAL_MONEY 实付金额 ,o.PRICE_REAL 商品采购价,o.COMMISSION 佣金金额 ,o.ACCOUNT_TYPE 充值账号类型 ,o.ACCOUNT_NUMBER 充值账号 ,o.PAY_TYPE 支付方式 ,
                o.PAY_CALL_TIME 支付唤起时间 ,o.PAY_TIME 支付时间,o.PAY_NODIFY_TIME 支付反馈时间,o.PAY_ORDER_NO 支付流水号,o.PAY_REMARK 支付备注,o.SHOP_ORDER_CALL_TIME 商家下单调用时间,
                o.SHOP_ORDER_FINISH_TIME 商家下单完成时间,o.SHOP_ORDER_STATUS_GET_TIME 商家下单状态获取时间,o.SHOP_ORDER_NO 商家订单号,o.SHOP_REMARK 商家备注,
                o.CREATE_TIME 创建时间,o.UPDATE_TIME 最后更新时间,
                case ORDER_STATUS WHEN 1 THEN '生成' WHEN 2 THEN '确认' END as 平台订单状态,
                case PAY_STATUS WHEN 0 THEN '0未调用' WHEN 1 THEN '1待支付'  WHEN 2 THEN '2已支付'  WHEN 3 THEN '3支付失败'  WHEN 4 THEN '4退款中'  WHEN 6 THEN '6退款成功' END as 支付状态,
                case SHOP_ORDER_STATUS WHEN 0 THEN '0未调用' WHEN 1 THEN '1未处理'  WHEN 2 THEN '2处理中'  WHEN 3 THEN '3成功'  WHEN 4 THEN '4失败'   END as 商家订单状态,
                case BRAND_TYPE WHEN 1 THEN '1直冲'  WHEN 2 THEN '2卡密'  END as 品牌类型,o.PHONE 手机号,o.PRODUCT_NAME 商品名称,
                de.CARD_NO 卡号,de.CARD_SECRET 密码,de.VALIDITY_TIME 有效期,de.CREATE_TIME 创建时间,
                case COUPON_TYPE WHEN 0 THEN '0二维码' WHEN 1 THEN '1条形码'  WHEN 2 THEN '2条形码和二维码'  WHEN 3 THEN '3卡券URL地址'  WHEN 4 THEN '4只包含密码'  WHEN 5 THEN '5卡号和密码'  END as 卡类型
                FROM interest_order o, interest_card_detail de
                where o.ORDER_NO=de.ORDER_NO
                ORDER BY o.CREATE_TIME desc;'''
        res,filed=self.selectre(tmpsql)
        # res=self.getresfloadtoint(res)
        # self.exportexcel(filed,res,"reportall")

        return res,filed,tmpsql
    # 查询权益中的商品
    def geproduct(self):
        tmpsql='''
                SELECT b.id '品牌id',b.BRAND_NAME '品牌名称',b.BRAND_DESC '品牌描述',
                CASE CATEGORY_ID
                    when 1 THEN '视频专区'
                    when 2 THEN '文娱专区'
                    when 3 THEN '旅游餐饮专区'
                        end as '类目',
                    b.BRAND_IMG '品牌',
                b.CREATE_TIME '创建时间',
                case b.BRAND_STATUS when 1 THEN '有效' WHEN 0 THEN '无效' END as '品牌状态',
                case b.BRAND_TYPE when 1 THEN '直冲' WHEN 2 THEN '卡密' END as '类型',
                b.PRIORITY '权重',
                b.SPECIAL_PROMPT '特别提示',
                p.id '商品id',
                p.PRODUCT_NAME '商品名称',
                p.SHOP_SHORT_NAME '商家名称',
                p.PRICE_PRI '商品原价',
                p.PRICE_CURRENT '商品现价',
                p.PRICE_REAL '商品采购价',
                p.PRODUCT_IMG '商品logo',
                p.NORM1 '规格值1',
                case p.PRODUCT_STATUS when 1 THEN '有效' WHEN 0 THEN '无效' END as '商品状态',
                p.PRIORITY '商品权重'
                -- p.REMARK '商品备注'
                FROM  interest_brand b,interest_product p
                where b.id=p.BRAND_ID
                   '''
        # headtr='''<tr><td>&nbsp;</td><td>&nbsp;</td><td>&nbsp;</td><td>&nbsp;</td><td>&nbsp;</td><td>&nbsp;</td><td>&nbsp;</td><td>&nbsp;</td><td>&nbsp;</td>
        #     <td align="center" colspan={0}>消耗</td><td align="center" colspan={1}>去除联动平台毛利</td><td align="center" colspan={2}>入口点击成本</td><td align="center" colspan={3}>去除联动毛利率</td></tr>'''.format(colspanx,colspanx,colspanx,colspanx)
        # print tmpsql
        res,filed=self.selectre(tmpsql)
        # res=self.getresfloadtoint(res)
        # self.exportexcel(filed,res,"reportall")

        return res,filed,tmpsql
    def getthirdproduct(self):
        tmpsql='''SELECT ib.ID 品牌id,ib.BRAND_NAME,ip.ID 商品id,ip.PRODUCT_ID 第三方商品id,
                case ip.PRODUCT_STATUS when 1 THEN '有效' WHEN 0 THEN '无效' END as '商品状态'
                FROM interest_product ip ,interest_brand ib
                where ib.ID=ip.BRAND_ID and ip.SHOP_SHORT_NAME='tq365'
                ORDER BY ip.PRODUCT_STATUS desc;'''
        res,filed=self.selectre(tmpsql)
        # filed=list(filed)
        filed.append(u'第三方库存')
        tmp=[]
        for i in res:
            num=self.getthirdprodcutstock(int(i[3]))
            i.append(num)
            tmp.append(i)
        return res,filed,tmpsql
    # 接口返回类似 {"msg":"库存数量：10","code":"0"} 数据，取得库存数量字段
    def getthirdprodcutstock(self,productid):
        url='http://221.122.127.206:18080/365tq/getNum'
        params={'productId': productid}
        re=r.get(url=url,params=params)
        num=re.json()['msg'][5:7]
        print type(int(num))
        return num


if __name__ == '__main__':
    # test=myreport(begintime='2020-04-1',endtime='2020-04-02',adzoneids='21')
    test=quanyi(env='test')
    # tmp,filed,tmpsql=test.geproduct()
    tmp,tmp1=test.getthirdproduct()
    print tmp,tmp1
    # print tmp
    # print get_date_list('2018-01-01','2018-02-28')
    # cwd = os.getcwd()
    # print(cwd)
    # wb=Workbook()
    # sheet=wb.active
    # wb1=load_workbook('reportall.xlsx')
    # w=wb1["Sheet"]
    # sheet.merge_cells('J1:L1')
    # sheet['J1']='消费消费'



