# encoding=utf-8
__author__ = 'aidinghua'
from openpyxl import  Workbook
import time
import datetime
import  sys
import re
from utils.db_info import *

# 有练换换	wx0a051787252f83fa
# 步数大联盟	wxe65c34b4ec242be
# 优质福利所	wx3c48ef7a45e89118
# 易购赚	wx0a051787252f83fa
# 遇见球球 	wxbb67c77aea9d76a8
# 开心游戏大乱斗	wx2fa65dcf3dbe5926
# 十一光年	wx42d12a5790960727
# 福礼惠公众号	wx4ef839c292cb41ad
# 我蹦我再蹦	wx85d4b50441231b98
# 弹球大乱斗	wx4aab7ee381936b54
# 手机弹幕小程序 wx3fe2d608967de425
class Media_report(object):

    def __init__(self,source_id,begin_date,end_date,env_value=False):

        self.dbinfo=DbOperations(env_value=env_value)
        self.begin_date=begin_date
        self.end_date=end_date
        self.source_id =source_id

    def begin_ym(self):

        begin_ym=self.begin_date[0:7]

        begin_ym1= begin_ym.replace('-','')
        return begin_ym1

    def end_ym(self):

        end_ym=self.end_date[0:7]
        end_ym1 = end_ym.replace('-','')
        return end_ym1




    def show_result(self):

        sql = '''
                  select
                  date_format(a.create_time, '%Y-%m-%d') 日期,

                  CASE
                  WHEN a.source_id='wx0a051787252f83fa' THEN '有练换换'
                  WHEN a.source_id='wxe65c34b4ec242be' THEN '步数大联盟'
                  WHEN a.source_id='wx3c48ef7a45e89118' THEN '优质福利所'
                  WHEN a.source_id='wx0a051787252f83fa' THEN '易购赚'
                  WHEN a.source_id='wxbb67c77aea9d76a8' THEN '遇见球球'
                  WHEN a.source_id='wx2fa65dcf3dbe5926' THEN '开心游戏大乱斗'
                  WHEN a.source_id='wx42d12a5790960727' THEN '十一光年'
                  WHEN a.source_id='wx4ef839c292cb41ad' THEN '福礼惠公众号'
                  WHEN a.source_id='wx85d4b50441231b98' THEN '我蹦我再蹦'
                  WHEN a.source_id='wx4aab7ee381936b54' THEN '弹球大乱斗'
                  WHEN a.source_id='1017757' THEN 'APP-DSP光速动力'
                  WHEN a.source_id='wx3fe2d608967de425' THEN '手机弹幕小程序'
                  ELSE '其他'
                  END  媒体 ,
                  count(a.open_id) 新用户数

                from
                  voyager.wx_user_info  a

                where
                a.source_id = '{}'
                and date_format(a.create_time,'%Y-%m-%d')>='{}' and date_format(a.create_time,'%Y-%m-%d')<='{}'
                group by DATE_FORMAT(a.create_time, '%Y-%m-%d'),媒体  order by  DATE_FORMAT(a.create_time, '%Y-%m-%d')  desc '''.format(self.source_id,self.begin_date,self.end_date)

        result =self.dbinfo.execute_sql(sql)

        if result<>():



            self.exportXlsadd(result)

        return result





##查询新增用户数
    def show_result2(self):
        sql = '''
                  select
                  date_format(a.create_time, '%Y-%m-%d') 日期,

                  CASE
                  WHEN a.source_id='wx0a051787252f83fa' THEN '有练换换'
                  WHEN a.source_id='wxe65c34b4ec242be' THEN '步数大联盟'
                  WHEN a.source_id='wx3c48ef7a45e89118' THEN '优质福利所'
                  WHEN a.source_id='wx0a051787252f83fa' THEN '易购赚'
                  WHEN a.source_id='wxbb67c77aea9d76a8' THEN '遇见球球'
                  WHEN a.source_id='wx2fa65dcf3dbe5926' THEN '开心游戏大乱斗'
                  WHEN a.source_id='wx42d12a5790960727' THEN '十一光年'
                  WHEN a.source_id='wx4ef839c292cb41ad' THEN '福礼惠公众号'
                  WHEN a.source_id='wx85d4b50441231b98' THEN '我蹦我再蹦'
                  WHEN a.source_id='wx4aab7ee381936b54' THEN '弹球大乱斗'
                  WHEN a.source_id='1017757' THEN 'APP-DSP光速动力'
                  WHEN a.source_id='wx3fe2d608967de425' THEN '手机弹幕小程序'
                  ELSE '其他'
                  END  媒体 ,
                  count(a.open_id) 新用户数

                from
                  voyager.wx_user_info  a

                where
                date_format(a.create_time,'%Y-%m-%d')>='{}' and date_format(a.create_time,'%Y-%m-%d')<='{}'
                group by DATE_FORMAT(a.create_time, '%Y-%m-%d'),媒体  order by  DATE_FORMAT(a.create_time, '%Y-%m-%d')  desc '''.format(self.begin_date,self.end_date)

        result =self.dbinfo.execute_sql(sql)



        if result<>():



            self.exportXlsadd(result)

        return  result



    def exportXlsadd(self,result):

        if len(str(result))>0:

            result = list(result)

            result.insert(0,('日期','媒体/活动','新用户数'))

            row = len(result)

            wb=Workbook()
            sheet=wb.active
            for i in range(row):

                sheet.append(result[i])


            wb.save('./static/result/adduser.xlsx')




##查询活跃用户数
    def liveuser(self):

        sql='''
              SELECT
              DATE_FORMAT(b.create_time, '%Y-%m-%d') 日期,

              CASE
                  WHEN a.source_id='wx0a051787252f83fa' THEN '有练换换'
                  WHEN a.source_id='wxe65c34b4ec242be' THEN '步数大联盟'
                  WHEN a.source_id='wx3c48ef7a45e89118' THEN '优质福利所'
                  WHEN a.source_id='wx0a051787252f83fa' THEN '易购赚'
                  WHEN a.source_id='wxbb67c77aea9d76a8' THEN '遇见球球'
                  WHEN a.source_id='wx2fa65dcf3dbe5926' THEN '开心游戏大乱斗'
                  WHEN a.source_id='wx42d12a5790960727' THEN '十一光年'
                  WHEN a.source_id='wx4ef839c292cb41ad' THEN '福礼惠公众号'
                  WHEN a.source_id='wx85d4b50441231b98' THEN '我蹦我再蹦'
                  WHEN a.source_id='wx4aab7ee381936b54' THEN '弹球大乱斗'
                  WHEN a.source_id='1017757' THEN 'APP-DSP光速动力'
                  WHEN a.source_id='wx3fe2d608967de425' THEN '手机弹幕小程序'
              ELSE '其他'
              END   媒体 ,
              COUNT(DISTINCT(b.open_id)) 活跃用户数
               FROM  voyager.wx_user_info a , voyagerlog.wx_user_log{} b  WHERE a.source_id ='{}' and  a.open_id=b.open_id  AND    DATE_FORMAT(b.create_time,'%Y-%m-%d')>='{}' AND DATE_FORMAT(b.create_time,'%Y-%m-%d')<='{}'
               GROUP BY  DATE_FORMAT(b.create_time,'%Y-%m-%d'),媒体  ORDER BY  DATE_FORMAT(b.create_time, '%Y-%m-%d')  DESC '''.format(self.begin_ym(),self.source_id,self.begin_date,self.end_date)
        result=self.dbinfo.execute_sql(sql)
        if result<>():


            self.exportXlslive(result)

        return result

    def liveuser2(self):

        sql='''
              SELECT
              DATE_FORMAT(b.create_time, '%Y-%m-%d') 日期,

              CASE
                  WHEN a.source_id='wx0a051787252f83fa' THEN '有练换换'
                  WHEN a.source_id='wxe65c34b4ec242be' THEN '步数大联盟'
                  WHEN a.source_id='wx3c48ef7a45e89118' THEN '优质福利所'
                  WHEN a.source_id='wx0a051787252f83fa' THEN '易购赚'
                  WHEN a.source_id='wxbb67c77aea9d76a8' THEN '遇见球球'
                  WHEN a.source_id='wx2fa65dcf3dbe5926' THEN '开心游戏大乱斗'
                  WHEN a.source_id='wx42d12a5790960727' THEN '十一光年'
                  WHEN a.source_id='wx4ef839c292cb41ad' THEN '福礼惠公众号'
                  WHEN a.source_id='wx85d4b50441231b98' THEN '我蹦我再蹦'
                  WHEN a.source_id='wx4aab7ee381936b54' THEN '弹球大乱斗'
                  WHEN a.source_id='1017757' THEN 'APP-DSP光速动力'
                  WHEN a.source_id='wx3fe2d608967de425' THEN '手机弹幕小程序'
              ELSE '其他'
              END   媒体 ,
              COUNT(DISTINCT(b.open_id)) 活跃用户数
               FROM  voyager.wx_user_info a , voyagerlog.wx_user_log{} b  WHERE  a.open_id=b.open_id  AND    DATE_FORMAT(b.create_time,'%Y-%m-%d')>='{}' AND DATE_FORMAT(b.create_time,'%Y-%m-%d')<='{}'
               GROUP BY  DATE_FORMAT(b.create_time,'%Y-%m-%d'),媒体  ORDER BY  DATE_FORMAT(b.create_time, '%Y-%m-%d')  DESC '''.format(self.begin_ym(),self.begin_date,self.end_date)

        result=self.dbinfo.execute_sql(sql)
        if result<>():



            self.exportXlslive(result)

        return result

    def exportXlslive(self,result):

        if len(str(result))>0:

            result=list(result)
            result.insert(0,('日期','媒体/活动','活跃用户数'))

            row=len(result)
            wb = Workbook()

            sheet = wb.active

            for i in range(row):
                sheet.append(result[i])

            wb.save('./static/result/liveuser.xlsx')


##查询授权用户数

    def authuser(self):

        sql='''
              SELECT
              DATE_FORMAT(a.create_time, '%Y-%m-%d') 日期,

              CASE
                  WHEN a.source_id='wx0a051787252f83fa' THEN '有练换换'
                  WHEN a.source_id='wxe65c34b4ec242be' THEN '步数大联盟'
                  WHEN a.source_id='wx3c48ef7a45e89118' THEN '优质福利所'
                  WHEN a.source_id='wx0a051787252f83fa' THEN '易购赚'
                  WHEN a.source_id='wxbb67c77aea9d76a8' THEN '遇见球球'
                  WHEN a.source_id='wx2fa65dcf3dbe5926' THEN '开心游戏大乱斗'
                  WHEN a.source_id='wx42d12a5790960727' THEN '十一光年'
                  WHEN a.source_id='wx4ef839c292cb41ad' THEN '福礼惠公众号'
                  WHEN a.source_id='wx85d4b50441231b98' THEN '我蹦我再蹦'
                  WHEN a.source_id='wx4aab7ee381936b54' THEN '弹球大乱斗'
                  WHEN a.source_id='1017757' THEN 'APP-DSP光速动力'
                  WHEN a.source_id='wx3fe2d608967de425' THEN '手机弹幕小程序'
              ELSE '其他'
              END  媒体 ,
              COUNT(a.open_id) 授权用户数

              FROM
              voyager.wx_user_info a
              where a.source_id = '{}'
              AND DATE_FORMAT(a.create_time,'%Y-%m-%d')>='{}' AND DATE_FORMAT(a.create_time,'%Y-%m-%d')<='{}'
              AND EXISTS (SELECT b.id FROM voyager.wx_user_authorize_log b WHERE b.open_id = a.open_id AND b.operate=1 AND b.type=1)
              AND EXISTS (SELECT b.id FROM voyager.wx_user_authorize_log b WHERE b.open_id = a.open_id AND b.operate=1 AND b.type=2)
              GROUP BY DATE_FORMAT(a.create_time, '%Y-%m-%d'),媒体  ORDER BY  DATE_FORMAT(a.create_time, '%Y-%m-%d') DESC '''.format(self.source_id,self.begin_date,self.end_date)

        result=self.dbinfo.execute_sql(sql)

        if result<>():



            self.exportXlsauth(result)

        return result


    def authuser2(self):

        sql='''
            SELECT
            DATE_FORMAT(a.create_time, '%Y-%m-%d') 日期,

            CASE
                  WHEN a.source_id='wx0a051787252f83fa' THEN '有练换换'
                  WHEN a.source_id='wxe65c34b4ec242be' THEN '步数大联盟'
                  WHEN a.source_id='wx3c48ef7a45e89118' THEN '优质福利所'
                  WHEN a.source_id='wx0a051787252f83fa' THEN '易购赚'
                  WHEN a.source_id='wxbb67c77aea9d76a8' THEN '遇见球球'
                  WHEN a.source_id='wx2fa65dcf3dbe5926' THEN '开心游戏大乱斗'
                  WHEN a.source_id='wx42d12a5790960727' THEN '十一光年'
                  WHEN a.source_id='wx4ef839c292cb41ad' THEN '福礼惠公众号'
                  WHEN a.source_id='wx85d4b50441231b98' THEN '我蹦我再蹦'
                  WHEN a.source_id='wx4aab7ee381936b54' THEN '弹球大乱斗'
                  WHEN a.source_id='1017757' THEN 'APP-DSP光速动力'
                  WHEN a.source_id='wx3fe2d608967de425' THEN '手机弹幕小程序'
            ELSE '其他'
            END  媒体 ,
            COUNT(a.open_id) 授权用户数
            FROM
            voyager.wx_user_info a
            WHERE
            DATE_FORMAT(a.create_time,'%Y-%m-%d')>='{}' AND DATE_FORMAT(a.create_time,'%Y-%m-%d')<='{}'
            AND EXISTS (
            SELECT b.id FROM voyager.wx_user_authorize_log b
            WHERE b.open_id = a.open_id AND b.operate=1 AND b.type=1) AND EXISTS (
            SELECT b.id FROM voyager.wx_user_authorize_log b
            WHERE b.open_id = a.open_id AND b.operate=1 AND b.type=2)
            GROUP BY DATE_FORMAT(a.create_time, '%Y-%m-%d'),媒体  ORDER BY  DATE_FORMAT(a.create_time, '%Y-%m-%d') DESC '''.format(self.begin_date,self.end_date)






        result=self.dbinfo.execute_sql(sql)

        if result<>():



            self.exportXlsauth(result)

        return result



    def exportXlsauth(self,result):

        if len(str(result))>0:

            result=list(result)
            result.insert(0,('日期','媒体/活动','授权用户数'))

            row=len(result)
            wb = Workbook()

            sheet = wb.active

            for i in range(row):
                sheet.append(result[i])

            wb.save('./static/result/authuser.xlsx')

    # def exportadduser(self,result):
    #
    #      if len(str(result)) > 0 :
    #
    #         result = list(result)
    #
    #         result.insert(0,('日期','媒体/活动','活跃用户数'))
    #
    #         row = len(result)
    #
    #         wb = Workbook()
    #         sheet = wb.active
    #
    #         for i in range(row):
    #
    #             sheet.append(result[i])
    #
    #         # wb.save("../../templates/ocpa_order.xlsx")
    #         wb.save("./static/result/ocpa_order.xlsx")






if __name__=='__main__':

    Me= Media_report('wx2fa65dcf3dbe5926','2019-02-01','2019-02-13')
    # print Me.show_result()
    # print Me.begin_ym()
    # print Me.end_ym()
    # print Me.liveuser()
    print Me.authuser()
